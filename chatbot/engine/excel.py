"""Build .xlsx workbooks for chatbot answers.

Uses openpyxl in write-only mode (bounded memory) — the same pattern the
existing platforms/reports.py export uses. Values are coerced so Decimals,
dates and tz-aware datetimes serialize cleanly.
"""

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Iterable, Sequence

from openpyxl import Workbook


def coerce(value: Any) -> Any:
    """Make a DB value safe for openpyxl / JSON."""
    if value is None:
        return ""
    if isinstance(value, Decimal):
        # Keep integers integer-looking, otherwise float.
        f = float(value)
        return int(f) if f.is_integer() else f
    if isinstance(value, datetime):
        # openpyxl rejects tz-aware datetimes.
        if value.tzinfo is not None:
            value = value.replace(tzinfo=None)
        return value
    if isinstance(value, date):
        return value
    if isinstance(value, (bytes, bytearray)):
        try:
            return value.decode("utf-8", "replace")
        except Exception:
            return str(value)
    return value


def _sanitize_sheet_title(title: str) -> str:
    bad = set('[]:*?/\\')
    clean = "".join(c for c in (title or "Sheet") if c not in bad).strip()
    return (clean or "Sheet")[:31]


def build_workbook(
    sheets: Sequence[tuple[str, Sequence[str], Iterable[Sequence[Any]]]],
    meta: Sequence[tuple[str, Any]] | None = None,
) -> tuple[bytes, int]:
    """Build a workbook from ``(title, columns, rows)`` tuples.

    Returns ``(bytes, total_data_rows)``.
    """
    wb = Workbook(write_only=True)
    total_rows = 0

    for title, columns, rows in sheets:
        ws = wb.create_sheet(_sanitize_sheet_title(title))
        ws.append([str(c) for c in columns])
        for row in rows:
            ws.append([coerce(v) for v in row])
            total_rows += 1

    if meta:
        ws2 = wb.create_sheet("About")
        ws2.append(["Field", "Value"])
        for key, val in meta:
            ws2.append([str(key), str(coerce(val))])

    if not wb.worksheets:  # never save an empty workbook
        ws = wb.create_sheet("Sheet1")
        ws.append(["No data"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), total_rows
