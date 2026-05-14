from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import connection, transaction

from openpyxl import load_workbook


MASTER_SHEET_FIELDS = (
    "format_sku_code",
    "product_name",
    "item",
    "format",
    "sku_sap_code",
    "sku_sap_name",
    "category",
    "sub_category",
    "case_pack",
    "per_unit",
    "item_head",
    "brand",
    "uom",
    "per_unit_value",
    "tax_rate",
    "category_head",
    "is_litre",
    "is_litre_oil",
    "packaging_cost",
)

MASTER_SHEET_UPDATE_FIELDS = (
    "product_name",
    "item",
    "sku_sap_code",
    "sku_sap_name",
    "category",
    "sub_category",
    "case_pack",
    "per_unit",
    "item_head",
    "brand",
    "uom",
    "per_unit_value",
    "tax_rate",
    "category_head",
    "is_litre",
    "is_litre_oil",
    "packaging_cost",
)


class Command(BaseCommand):
    help = "Seed public.master_sheet and FC master from the transformed Amazon PO workbook."

    def add_arguments(self, parser):
        parser.add_argument("workbook", help="Path to Amazon PO.xlsx reference workbook")

    def handle(self, *args, **options):
        path = Path(options["workbook"])
        if not path.exists():
            raise CommandError(f"Workbook not found: {path}")

        workbook = load_workbook(path, data_only=True, read_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = list(next(rows, []))
        index = {str(header).strip(): idx for idx, header in enumerate(headers) if header}

        required = {"SKU Code", "SKU Name", "Fulfillment Center"}
        missing = sorted(required - set(index))
        if missing:
            raise CommandError(f"Missing required reference columns: {', '.join(missing)}")

        product_seen: set[str] = set()
        fc_seen: set[str] = set()
        products = []
        fcs = []
        for row in rows:
            get = lambda name: row[index[name]] if name in index and index[name] < len(row) else None
            format_sku_code = _text(get("SKU Code"))
            product_name = _text(get("SKU Name"))
            if format_sku_code or product_name:
                key = (format_sku_code or product_name or "").upper()
                if key not in product_seen:
                    product_seen.add(key)
                    per_unit = _text(get("PER LTR UNIT"))
                    uom = _text(get("Unit of Measure"))
                    is_litre = _is_litre(per_unit, uom)
                    products.append(
                        {
                            "format_sku_code": format_sku_code,
                            "product_name": product_name,
                            "item": _text(get("ITEM")),
                            "format": "AMAZON",
                            "sku_sap_code": _text(get("SAP SKU Code")),
                            "sku_sap_name": _text(get("SAP SKU NAME")),
                            "category": _text(get("Category")),
                            "sub_category": _text(get("Sub Category")),
                            "case_pack": _int(get("Case Pack")),
                            "per_unit": per_unit,
                            "item_head": _text(get("Item Head")),
                            "brand": _text(get("Brand")),
                            "uom": uom,
                            "per_unit_value": _decimal(get("Per Liter")),
                            "tax_rate": _decimal(
                                _first_value(get, "TAX RATE", "Tax Rate", "TAX", "Tax")
                            ),
                            "category_head": _text(get("Category Head")),
                            "is_litre": "Y" if is_litre else "N",
                            "is_litre_oil": "Y" if is_litre and _text(get("Category Head")) == "OIL" else "N",
                            "packaging_cost": None,
                        }
                    )

            fc_code = _text(get("Fulfillment Center"))
            if fc_code and fc_code not in fc_seen:
                fc_seen.add(fc_code)
                fcs.append(
                    {
                        "fc_code": fc_code,
                        "city": _text(get("City")),
                        "state": _text(get("State")),
                    }
                )
        workbook.close()

        with transaction.atomic():
            with connection.cursor() as cur:
                for product in products:
                    _upsert_product(cur, product)
                for fc in fcs:
                    cur.execute(
                        """
                        INSERT INTO master.fc_master (fc_code, city, state, is_active, updated_at)
                        VALUES (%s, %s, %s, true, now())
                        ON CONFLICT (fc_code) DO UPDATE SET
                            city = EXCLUDED.city,
                            state = EXCLUDED.state,
                            is_active = true,
                            updated_at = now()
                        """,
                        [fc["fc_code"], fc["city"], fc["state"]],
                    )

        self.stdout.write(
            self.style.SUCCESS(
                f"Imported/updated {len(products)} master_sheet rows and {len(fcs)} FC rows."
            )
        )


def _upsert_product(cur, product: dict[str, Any]) -> None:
    values = [product[field] for field in MASTER_SHEET_FIELDS]
    set_sql = ", ".join(f"{field} = %s" for field in MASTER_SHEET_UPDATE_FIELDS)
    cur.execute(
        f"""
        UPDATE public.master_sheet
           SET {set_sql},
               format = 'AMAZON'
         WHERE UPPER(COALESCE(format, '')) = 'AMAZON'
           AND UPPER(TRIM(format_sku_code::text)) = UPPER(TRIM(%s))
        RETURNING format_sku_code
        """,
        [
            *[product[field] for field in MASTER_SHEET_UPDATE_FIELDS],
            product["format_sku_code"],
        ],
    )
    if cur.fetchone():
        return

    columns = ", ".join(MASTER_SHEET_FIELDS)
    placeholders = ", ".join(["%s"] * len(MASTER_SHEET_FIELDS))
    cur.execute(
        f"""
        INSERT INTO public.master_sheet ({columns})
        VALUES ({placeholders})
        """,
        values,
    )


def _text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal):
        return str(value.normalize())
    text = str(value).strip()
    return text or None


def _decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None


def _first_value(get, *names: str) -> Any:
    for name in names:
        value = get(name)
        if value is not None and value != "":
            return value
    return None


def _int(value: Any) -> int | None:
    decimal_value = _decimal(value)
    if decimal_value is None:
        return None
    return int(decimal_value)


def _is_litre(per_unit: str | None, uom: str | None) -> bool:
    text = f"{per_unit or ''} {uom or ''}".upper()
    return any(token in text for token in ("LTR", "LITRE", "ML"))
