from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import uuid
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.conf import settings
from django.db import connection, transaction
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.exceptions import PermissionDenied
from rest_framework.response import Response

from accounts.permissions import can_access_platform, require

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - exercised only when dependency missing
    load_workbook = None


MAX_UPLOAD_SIZE = 20 * 1024 * 1024
ALLOWED_EXTENSIONS = {".csv", ".xlsx"}
UPLOAD_DIR = Path(settings.BASE_DIR) / "uploaded_files" / "amazon"


@dataclass(frozen=True)
class ReportConfig:
    report_type: str
    label: str
    main_table_name: str
    raw_file_name: str
    staging_table_sql: str
    final_table_sql: str
    aliases: dict[str, str]
    staging_columns: tuple[str, ...]
    required_columns: tuple[str, ...]
    numeric_fields: tuple[str, ...] = ()
    date_fields: tuple[str, ...] = ()
    datetime_fields: tuple[str, ...] = ()
    text_fields: tuple[str, ...] = ()


AMAZON_PO_ALIASES = {
    "PO": "po_number",
    "Vendor code": "vendor_code",
    "Order date": "order_date",
    "Status": "status",
    "Product name": "product_name",
    "ASIN": "asin",
    "External ID type": "external_id_type",
    "External ID": "external_id",
    "Model number": "model_number",
    "Merchant SKU": "merchant_sku",
    "Catalog number": "catalog_number",
    "Availability": "availability",
    "Requested quantity": "requested_quantity",
    "Accepted quantity": "accepted_quantity",
    "ASN quantity": "asn_quantity",
    "Received quantity": "received_quantity",
    "Cancelled quantity": "cancelled_quantity",
    "Remaining quantity": "remaining_quantity",
    "Ship-to location": "ship_to_location",
    "Ship to location": "ship_to_location",
    "Window start": "window_start",
    "Window end": "window_end",
    "Case size": "case_size",
    "Cost": "cost",
    "Currency": "currency",
    "Total requested cost": "total_requested_cost",
    "Total accepted cost": "total_accepted_cost",
    "Total received cost": "total_received_cost",
    "Total cancelled cost": "total_cancelled_cost",
    "Expected date": "expected_date",
    "Freight terms": "freight_terms",
    "Consolidation ID": "consolidation_id",
    "Cancellation deadline": "cancellation_deadline",
}

APPOINTMENT_ALIASES = {
    "Appointment Id": "appointment_id",
    "Appointment ID": "appointment_id",
    "Status": "status",
    "Appointment Time": "appointment_time",
    "Creation Date": "creation_date",
    "POs": "pos",
    "Destination FC": "destination_fc",
    "PRO": "pro",
}

AMAZON_NUMERIC_FIELDS = (
    "requested_quantity",
    "accepted_quantity",
    "asn_quantity",
    "received_quantity",
    "cancelled_quantity",
    "remaining_quantity",
    "case_size",
    "cost",
    "total_requested_cost",
    "total_accepted_cost",
    "total_received_cost",
    "total_cancelled_cost",
)

REPORTS: dict[str, ReportConfig] = {
    "AMAZON_PO": ReportConfig(
        report_type="AMAZON_PO",
        label="Amazon PO",
        main_table_name="Amazon PO",
        raw_file_name="amazon data",
        staging_table_sql='staging."amazon data"',
        final_table_sql='reporting."Amazon PO"',
        aliases=AMAZON_PO_ALIASES,
        staging_columns=(
            "po_number",
            "vendor_code",
            "order_date",
            "status",
            "product_name",
            "asin",
            "external_id_type",
            "external_id",
            "model_number",
            "merchant_sku",
            "catalog_number",
            "availability",
            "requested_quantity",
            "accepted_quantity",
            "asn_quantity",
            "received_quantity",
            "cancelled_quantity",
            "remaining_quantity",
            "ship_to_location",
            "window_start",
            "window_end",
            "case_size",
            "cost",
            "currency",
            "total_requested_cost",
            "total_accepted_cost",
            "total_received_cost",
            "total_cancelled_cost",
            "expected_date",
            "freight_terms",
            "consolidation_id",
            "cancellation_deadline",
        ),
        required_columns=("po_number", "order_date", "asin", "ship_to_location"),
        numeric_fields=AMAZON_NUMERIC_FIELDS,
        date_fields=(
            "order_date",
            "window_start",
            "window_end",
            "expected_date",
            "cancellation_deadline",
        ),
    ),
    "APPOINTMENT": ReportConfig(
        report_type="APPOINTMENT",
        label="appointment",
        main_table_name="appointment",
        raw_file_name="appointment data",
        staging_table_sql='staging."appointment data"',
        final_table_sql='reporting."appointment"',
        aliases=APPOINTMENT_ALIASES,
        staging_columns=(
            "appointment_id",
            "status",
            "appointment_time",
            "creation_date",
            "pos",
            "destination_fc",
            "pro",
        ),
        required_columns=("appointment_id", "appointment_time", "destination_fc"),
        date_fields=("creation_date",),
        datetime_fields=("appointment_time",),
    ),
}

MAIN_TO_REPORT_TYPE = {cfg.main_table_name: key for key, cfg in REPORTS.items()}


def normalize_header(value: Any) -> str:
    text = str(value or "").replace("\ufeff", "")
    text = re.sub(r"[\r\n\t]+", " ", text)
    text = re.sub(r"[\u2010-\u2015\u2212]+", "-", text)
    text = text.strip().lower()
    text = re.sub(r"\s+", " ", text)
    text = text.strip(" .,:;|/\\()[]{}")
    return text


def _alias_map(config: ReportConfig) -> dict[str, str]:
    return {normalize_header(src): dest for src, dest in config.aliases.items()}


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _parse_numeric(value: Any) -> tuple[Decimal | None, str | None]:
    if value is None:
        return None, None
    if isinstance(value, Decimal):
        return value, None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return Decimal(str(value)), None
    text = str(value).strip()
    if text in {"", "-"}:
        return None, None
    text = re.sub(r"[₹$€£,\s]", "", text)
    text = text.rstrip("%")
    if text in {"", "-"}:
        return None, None
    try:
        return Decimal(text), None
    except InvalidOperation:
        return None, f"Invalid numeric value: {value}"


def _excel_serial_to_date(value: int | float) -> datetime | None:
    try:
        if value <= 0:
            return None
        return datetime(1899, 12, 30) + timedelta(days=float(value))
    except (OverflowError, ValueError):
        return None


def _parse_temporal(
    value: Any,
    *,
    want_datetime: bool,
    prefer_month_first: bool = False,
) -> tuple[date | datetime | None, str | None]:
    if value is None:
        return None, None
    if isinstance(value, datetime):
        return (value if want_datetime else value.date()), None
    if isinstance(value, date):
        return (datetime.combine(value, time.min) if want_datetime else value), None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        converted = _excel_serial_to_date(value)
        if converted:
            return (converted if want_datetime else converted.date()), None
    text = str(value).strip()
    if text in {"", "-"}:
        return None, None
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"\s+(IST|UTC|GMT)$", "", text, flags=re.IGNORECASE).strip()
    slash_date = re.match(r"^(\d{1,2})/\d{1,2}/\d{2,4}(?:\s|$)", text)
    if prefer_month_first and slash_date and not slash_date.group(1).startswith("0"):
        preferred_formats = (
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%Y %H:%M",
            "%m/%d/%Y %I:%M:%S %p",
            "%m/%d/%Y %I:%M %p",
            "%m/%d/%y %H:%M:%S",
            "%m/%d/%y %H:%M",
            "%m/%d/%y %I:%M:%S %p",
            "%m/%d/%y %I:%M %p",
            "%m/%d/%Y",
            "%m/%d/%y",
        )
        for fmt in preferred_formats:
            try:
                parsed = datetime.strptime(text, fmt)
                return (parsed if want_datetime else parsed.date()), None
            except ValueError:
                pass
    formats = (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y %I:%M:%S %p",
        "%d-%m-%Y %I:%M %p",
        "%d-%m-%y %I:%M:%S %p",
        "%d-%m-%y %I:%M %p",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %I:%M:%S %p",
        "%d/%m/%Y %I:%M %p",
        "%d/%m/%y %I:%M:%S %p",
        "%d/%m/%y %I:%M %p",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y %I:%M %p",
        "%m/%d/%y %I:%M:%S %p",
        "%m/%d/%y %I:%M %p",
        "%d-%b-%Y %H:%M:%S",
        "%d-%b-%Y %H:%M",
        "%d-%b-%Y %I:%M:%S %p",
        "%d-%b-%Y %I:%M %p",
        "%d-%B-%Y %H:%M:%S",
        "%d-%B-%Y %H:%M",
        "%d-%B-%Y %I:%M:%S %p",
        "%d-%B-%Y %I:%M %p",
        "%d-%b-%y %H:%M:%S",
        "%d-%b-%y %H:%M",
        "%d-%b-%y %I:%M:%S %p",
        "%d-%b-%y %I:%M %p",
        "%d-%B-%y %H:%M:%S",
        "%d-%B-%y %H:%M",
        "%d-%B-%y %I:%M:%S %p",
        "%d-%B-%y %I:%M %p",
        "%d %b %Y %H:%M:%S",
        "%d %b %Y %H:%M",
        "%d %b %Y %I:%M:%S %p",
        "%d %b %Y %I:%M %p",
        "%d %B %Y %H:%M:%S",
        "%d %B %Y %H:%M",
        "%d %B %Y %I:%M:%S %p",
        "%d %B %Y %I:%M %p",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d-%m-%y",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%d.%m.%Y",
        "%d-%b-%Y",
        "%d-%b-%y",
        "%d-%B-%Y",
        "%d-%B-%y",
        "%d %b %Y",
        "%d %b %y",
        "%d %B %Y",
        "%d %B %y",
        "%Y/%m/%d",
    )
    for fmt in formats:
        try:
            parsed = datetime.strptime(text, fmt)
            return (parsed if want_datetime else parsed.date()), None
        except ValueError:
            pass
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        if parsed.tzinfo:
            parsed = parsed.astimezone(timezone.get_current_timezone()).replace(tzinfo=None)
        return (parsed if want_datetime else parsed.date()), None
    except ValueError:
        return None, f"Invalid {'datetime' if want_datetime else 'date'} value: {value}"


def _normalize_text_value(config: ReportConfig, field: str, value: Any) -> str | None:
    if _is_blank(value):
        return None
    if isinstance(value, Decimal):
        text = format(value, "f")
    elif isinstance(value, int) and not isinstance(value, bool):
        text = str(value)
    elif isinstance(value, float) and not isinstance(value, bool):
        text = str(int(value)) if value.is_integer() else str(value)
    else:
        text = str(value).strip()

    if config.report_type == "APPOINTMENT" and field == "appointment_id":
        cleaned = text.replace(",", "").strip()
        try:
            parsed = Decimal(cleaned)
            if parsed == parsed.to_integral_value():
                return format(parsed.quantize(Decimal("1")), "f")
        except InvalidOperation:
            pass
    return text


def _read_csv(content: bytes) -> list[list[Any]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    first_data_line = next((line for line in text.splitlines() if line.strip()), "")
    if "\t" in first_data_line:
        return [row for row in csv.reader(io.StringIO(text), delimiter="\t")]
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel
    return [row for row in csv.reader(io.StringIO(text), dialect)]


def _read_xlsx(content: bytes) -> list[list[Any]]:
    if load_workbook is None:
        raise ValueError("XLSX uploads require openpyxl. Install backend requirements first.")
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    return rows


def parse_uploaded_file(
    *,
    config: ReportConfig,
    content: bytes,
    extension: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], int]:
    table = _read_csv(content) if extension == ".csv" else _read_xlsx(content)
    table = [row for row in table if any(not _is_blank(cell) for cell in row)]
    if not table:
        return [], [
            {
                "row_number": None,
                "field_name": None,
                "error_type": "empty_file",
                "error_message": "The uploaded file does not contain any rows.",
                "severity": "error",
            }
        ], 0

    headers = [normalize_header(cell) for cell in table[0]]
    aliases = _alias_map(config)
    header_map: dict[int, str] = {}
    mapped_fields = set()
    issues: list[dict[str, Any]] = []
    for idx, header in enumerate(headers):
        field = aliases.get(header)
        if field and field not in mapped_fields:
            header_map[idx] = field
            mapped_fields.add(field)

    for field in config.required_columns:
        if field not in mapped_fields:
            issues.append(
                {
                    "row_number": None,
                    "field_name": field,
                    "error_type": "missing_required_column",
                    "error_message": f"Required column '{field}' is missing.",
                    "severity": "error",
                }
            )
    parsed_rows: list[dict[str, Any]] = []
    for row_idx, raw_row in enumerate(table[1:], start=2):
        row = {col: None for col in config.staging_columns}
        row["raw_row_number"] = row_idx
        for idx, field in header_map.items():
            value = raw_row[idx] if idx < len(raw_row) else None
            if field in config.numeric_fields:
                parsed, err = _parse_numeric(value)
                row[field] = parsed
                if err:
                    issues.append(
                        {
                            "row_number": row_idx,
                            "field_name": field,
                            "error_type": "invalid_numeric",
                            "error_message": err,
                            "severity": "error",
                        }
                    )
            elif field in config.datetime_fields:
                parsed, err = _parse_temporal(value, want_datetime=True)
                row[field] = parsed
                if err:
                    issues.append(
                        {
                            "row_number": row_idx,
                            "field_name": field,
                            "error_type": "invalid_date",
                            "error_message": err,
                            "severity": "error",
                        }
                    )
            elif field in config.date_fields:
                parsed, err = _parse_temporal(
                    value,
                    want_datetime=False,
                    prefer_month_first=(
                        config.report_type == "APPOINTMENT" and field == "creation_date"
                    ),
                )
                row[field] = parsed
                if err:
                    issues.append(
                        {
                            "row_number": row_idx,
                            "field_name": field,
                            "error_type": "invalid_date",
                            "error_message": err,
                            "severity": "error",
                        }
                    )
            else:
                row[field] = _normalize_text_value(config, field, value)

        for field in config.required_columns:
            if _is_blank(row.get(field)):
                issues.append(
                    {
                        "row_number": row_idx,
                        "field_name": field,
                        "error_type": "missing_required_value",
                        "error_message": f"Required value '{field}' is blank.",
                        "severity": "error",
                    }
                )
        parsed_rows.append(row)

    return parsed_rows, issues, max(len(table) - 1, 0)


def _safe_filename(name: str) -> str:
    stem = Path(name or "upload").stem[:80]
    suffix = Path(name or "").suffix.lower()
    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("._-") or "upload"
    return f"{timezone.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:10]}_{stem}{suffix}"


def _store_file(content: bytes, original_name: str) -> str:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    target = UPLOAD_DIR / _safe_filename(original_name)
    target.write_bytes(content)
    return str(target)


def _insert_upload_file(
    cur,
    *,
    config: ReportConfig,
    original_file_name: str,
    stored_file_path: str | None,
    file_hash: str | None,
    file_extension: str,
    uploaded_by: str,
    status_value: str,
    metadata: dict[str, Any],
) -> int:
    cur.execute(
        """
        INSERT INTO raw.upload_file (
            main_table_name, raw_file_name, original_file_name, stored_file_path,
            file_hash, file_extension, uploaded_by, status, metadata
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb)
        RETURNING upload_id
        """,
        [
            config.main_table_name,
            config.raw_file_name,
            original_file_name,
            stored_file_path,
            file_hash,
            file_extension.lstrip("."),
            uploaded_by,
            status_value,
            json.dumps(metadata),
        ],
    )
    return int(cur.fetchone()[0])


def _update_upload_file(
    cur,
    *,
    upload_id: int,
    status_value: str,
    row_count: int,
    error_count: int,
    warning_count: int,
    metadata: dict[str, Any] | None = None,
) -> None:
    cur.execute(
        """
        UPDATE raw.upload_file
           SET status = %s,
               row_count = %s,
               error_count = %s,
               warning_count = %s,
               metadata = COALESCE(metadata, '{}'::jsonb) || %s::jsonb,
               updated_at = now()
         WHERE upload_id = %s
        """,
        [
            status_value,
            row_count,
            error_count,
            warning_count,
            json.dumps(metadata or {}),
            upload_id,
        ],
    )


def _insert_validation_issues(
    cur,
    *,
    upload_id: int,
    config: ReportConfig,
    issues: list[dict[str, Any]],
) -> None:
    if not issues:
        return
    rows = [
        [
            upload_id,
            config.main_table_name,
            config.raw_file_name,
            issue.get("row_number"),
            issue.get("field_name"),
            issue.get("error_type") or "validation",
            issue.get("error_message") or "Validation issue",
            issue.get("severity") or "error",
        ]
        for issue in issues
    ]
    cur.executemany(
        """
        INSERT INTO quality.validation_error (
            upload_id, main_table_name, raw_file_name, row_number, field_name,
            error_type, error_message, severity
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """,
        rows,
    )


def _upsert_summary(
    cur,
    *,
    upload_id: int,
    config: ReportConfig,
    total_rows: int,
    valid_rows: int,
    error_rows: int,
    warning_rows: int,
    inserted: int,
    updated: int,
    status_value: str,
) -> None:
    cur.execute(
        """
        INSERT INTO quality.upload_validation_summary (
            upload_id, main_table_name, raw_file_name, total_rows, valid_rows,
            error_rows, warning_rows, final_inserted_rows, final_updated_rows, status
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (upload_id) DO UPDATE SET
            total_rows = EXCLUDED.total_rows,
            valid_rows = EXCLUDED.valid_rows,
            error_rows = EXCLUDED.error_rows,
            warning_rows = EXCLUDED.warning_rows,
            final_inserted_rows = EXCLUDED.final_inserted_rows,
            final_updated_rows = EXCLUDED.final_updated_rows,
            status = EXCLUDED.status
        """,
        [
            upload_id,
            config.main_table_name,
            config.raw_file_name,
            total_rows,
            valid_rows,
            error_rows,
            warning_rows,
            inserted,
            updated,
            status_value,
        ],
    )


def _insert_staging_rows(
    cur,
    *,
    upload_id: int,
    config: ReportConfig,
    rows: list[dict[str, Any]],
) -> int:
    if not rows:
        return 0
    columns = ("upload_id", "raw_row_number", *config.staging_columns)
    quoted_columns = ", ".join(f'"{col}"' for col in columns)
    placeholders = ", ".join(["%s"] * len(columns))
    values = [
        [upload_id, row.get("raw_row_number"), *[row.get(col) for col in config.staging_columns]]
        for row in rows
    ]
    cur.executemany(
        f"INSERT INTO {config.staging_table_sql} ({quoted_columns}) VALUES ({placeholders})",
        values,
    )
    return len(values)


def _master_warning_for_amazon(cur, row: dict[str, Any]) -> list[dict[str, Any]]:
    warnings: list[dict[str, Any]] = []
    cur.execute(
        """
        SELECT format_sku_code, case_pack,
               COALESCE(
                   NULLIF(per_unit_value::numeric, 0),
                   CASE
                       WHEN NULLIF(TRIM(per_unit::text), '') IS NULL THEN NULL
                       WHEN substring(per_unit::text from '([0-9]+(?:\.[0-9]+)?)') IS NULL THEN NULL
                       WHEN UPPER(COALESCE(uom, '')) IN ('ML', 'MLS')
                            OR UPPER(per_unit::text) LIKE '%%ML%%'
                            THEN substring(per_unit::text from '([0-9]+(?:\.[0-9]+)?)')::numeric / 1000
                       WHEN UPPER(COALESCE(uom, '')) IN ('LTR', 'LITRE', 'LITRES')
                            OR UPPER(per_unit::text) LIKE '%%LTR%%'
                            OR UPPER(per_unit::text) LIKE '%%LITRE%%'
                            THEN substring(per_unit::text from '([0-9]+(?:\.[0-9]+)?)')::numeric
                       ELSE NULL
                   END
               ) AS per_unit_value,
               per_unit,
               uom, sku_sap_name, sku_sap_code
          FROM public.master_sheet
         WHERE (
                (NULLIF(%s, '') IS NOT NULL AND UPPER(TRIM(format_sku_code::text)) = UPPER(TRIM(%s)))
             OR (NULLIF(%s, '') IS NOT NULL AND UPPER(TRIM(format_sku_code::text)) = UPPER(TRIM(%s)))
             OR (NULLIF(%s, '') IS NOT NULL AND UPPER(TRIM(item::text)) = UPPER(TRIM(%s)))
             OR (NULLIF(%s, '') IS NOT NULL AND LOWER(TRIM(product_name::text)) = LOWER(TRIM(%s)))
           )
         ORDER BY CASE
             WHEN UPPER(COALESCE(format, '')) = 'AMAZON'
                  AND NULLIF(%s, '') IS NOT NULL
                  AND UPPER(TRIM(format_sku_code::text)) = UPPER(TRIM(%s)) THEN 1
             WHEN UPPER(COALESCE(format, '')) = 'AMAZON'
                  AND NULLIF(%s, '') IS NOT NULL
                  AND UPPER(TRIM(format_sku_code::text)) = UPPER(TRIM(%s)) THEN 2
             WHEN UPPER(COALESCE(format, '')) = 'AMAZON'
                  AND NULLIF(%s, '') IS NOT NULL
                  AND UPPER(TRIM(item::text)) = UPPER(TRIM(%s)) THEN 3
             WHEN UPPER(COALESCE(format, '')) = 'AMAZON'
                  AND NULLIF(%s, '') IS NOT NULL
                  AND LOWER(TRIM(product_name::text)) = LOWER(TRIM(%s)) THEN 4
             ELSE 5
         END
         LIMIT 1
        """,
        [
            row.get("asin") or "",
            row.get("asin") or "",
            row.get("external_id") or "",
            row.get("external_id") or "",
            row.get("merchant_sku") or "",
            row.get("merchant_sku") or "",
            row.get("product_name") or "",
            row.get("product_name") or "",
            row.get("asin") or "",
            row.get("asin") or "",
            row.get("external_id") or "",
            row.get("external_id") or "",
            row.get("merchant_sku") or "",
            row.get("merchant_sku") or "",
            row.get("product_name") or "",
            row.get("product_name") or "",
        ],
    )
    product = cur.fetchone()
    row_number = row.get("raw_row_number")
    if not product:
        warnings.append(
            {
                "row_number": row_number,
                "field_name": "master_sheet",
                "error_type": "master_sheet_missing",
                "error_message": "Master sheet mapping missing.",
                "severity": "warning",
            }
        )
    else:
        has_pack_reference = bool(product[5] or product[6])
        case_pack = product[1] or row.get("case_size")
        unit_text = f"{product[3] or ''} {product[4] or ''}".upper()
        is_litre_product = any(token in unit_text for token in ("LTR", "LITRE", "ML"))
        if case_pack in (None, 0) and has_pack_reference:
            warnings.append(
                {
                    "row_number": row_number,
                    "field_name": "case_pack",
                    "error_type": "case_pack_missing",
                    "error_message": "case_pack missing in master_sheet.",
                    "severity": "warning",
                }
            )
        if product[2] in (None, 0) and is_litre_product:
            warnings.append(
                {
                    "row_number": row_number,
                    "field_name": "per_unit_value",
                    "error_type": "per_unit_value_missing",
                    "error_message": "per_unit_value missing in master_sheet.",
                    "severity": "warning",
                }
            )

    cur.execute(
        """
        SELECT 1
          FROM master.fc_master
         WHERE is_active = true AND fc_code = %s
         LIMIT 1
        """,
        [row.get("ship_to_location")],
    )
    if row.get("ship_to_location") and cur.fetchone() is None:
        warnings.append(
            {
                "row_number": row_number,
                "field_name": "ship_to_location",
                "error_type": "fc_master_missing",
                "error_message": "FC master mapping missing.",
                "severity": "warning",
            }
        )
    return warnings


def _add_business_warnings(
    cur,
    *,
    config: ReportConfig,
    upload_id: int,
) -> list[dict[str, Any]]:
    if config.report_type != "AMAZON_PO":
        return []
    cur.execute(
        """
        WITH src AS (
            SELECT raw_row_number, external_id, asin, merchant_sku, product_name,
                   ship_to_location, case_size
              FROM staging."amazon data"
             WHERE upload_id = %s
        ),
        matched AS (
            SELECT s.*,
                   pm.format_sku_code,
                   COALESCE(pm.case_pack, NULLIF(s.case_size, 0)) AS case_pack,
                   pm.per_unit_value,
                   pm.per_unit,
                   pm.uom,
                   pm.sku_sap_name,
                   pm.sku_sap_code,
                   fc.fc_id
              FROM src s
              LEFT JOIN LATERAL (
                  SELECT format_sku_code, case_pack,
                         COALESCE(
                             NULLIF(per_unit_value::numeric, 0),
                             CASE
                                 WHEN NULLIF(TRIM(per_unit::text), '') IS NULL THEN NULL
                                 WHEN substring(per_unit::text from '([0-9]+(?:\.[0-9]+)?)') IS NULL THEN NULL
                                 WHEN UPPER(COALESCE(uom, '')) IN ('ML', 'MLS')
                                      OR UPPER(per_unit::text) LIKE '%%ML%%'
                                      THEN substring(per_unit::text from '([0-9]+(?:\.[0-9]+)?)')::numeric / 1000
                                 WHEN UPPER(COALESCE(uom, '')) IN ('LTR', 'LITRE', 'LITRES')
                                      OR UPPER(per_unit::text) LIKE '%%LTR%%'
                                      OR UPPER(per_unit::text) LIKE '%%LITRE%%'
                                      THEN substring(per_unit::text from '([0-9]+(?:\.[0-9]+)?)')::numeric
                                 ELSE NULL
                             END
                         ) AS per_unit_value,
                         per_unit,
                         uom, sku_sap_name, sku_sap_code
                    FROM public.master_sheet pm
                   WHERE (
                          (NULLIF(s.asin, '') IS NOT NULL AND UPPER(TRIM(pm.format_sku_code::text)) = UPPER(TRIM(s.asin::text)))
                       OR (NULLIF(s.external_id, '') IS NOT NULL AND UPPER(TRIM(pm.format_sku_code::text)) = UPPER(TRIM(s.external_id::text)))
                       OR (NULLIF(s.merchant_sku, '') IS NOT NULL AND UPPER(TRIM(pm.item::text)) = UPPER(TRIM(s.merchant_sku::text)))
                       OR (NULLIF(s.product_name, '') IS NOT NULL AND LOWER(TRIM(pm.product_name::text)) = LOWER(TRIM(s.product_name::text)))
                     )
                   ORDER BY CASE
                       WHEN UPPER(COALESCE(pm.format, '')) = 'AMAZON'
                            AND NULLIF(s.asin, '') IS NOT NULL
                            AND UPPER(TRIM(pm.format_sku_code::text)) = UPPER(TRIM(s.asin::text)) THEN 1
                       WHEN UPPER(COALESCE(pm.format, '')) = 'AMAZON'
                            AND NULLIF(s.external_id, '') IS NOT NULL
                            AND UPPER(TRIM(pm.format_sku_code::text)) = UPPER(TRIM(s.external_id::text)) THEN 2
                       WHEN UPPER(COALESCE(pm.format, '')) = 'AMAZON'
                            AND NULLIF(s.merchant_sku, '') IS NOT NULL
                            AND UPPER(TRIM(pm.item::text)) = UPPER(TRIM(s.merchant_sku::text)) THEN 3
                       WHEN UPPER(COALESCE(pm.format, '')) = 'AMAZON'
                            AND NULLIF(s.product_name, '') IS NOT NULL
                            AND LOWER(TRIM(pm.product_name::text)) = LOWER(TRIM(s.product_name::text)) THEN 4
                       ELSE 5
                   END
                   LIMIT 1
              ) pm ON true
              LEFT JOIN master.fc_master fc
                ON fc.is_active = true AND fc.fc_code = s.ship_to_location
        )
        SELECT raw_row_number, field_name, error_type, error_message
          FROM (
                SELECT raw_row_number,
                       'master_sheet' AS field_name,
                       'master_sheet_missing' AS error_type,
                       'Master sheet mapping missing.' AS error_message
                  FROM matched
                 WHERE format_sku_code IS NULL

                UNION ALL

                SELECT raw_row_number,
                       'ship_to_location' AS field_name,
                       'fc_master_missing' AS error_type,
                       'FC master mapping missing.' AS error_message
                  FROM matched
                 WHERE NULLIF(TRIM(ship_to_location), '') IS NOT NULL
                   AND fc_id IS NULL

                UNION ALL

                SELECT raw_row_number,
                       'case_pack' AS field_name,
                       'case_pack_missing' AS error_type,
                       'case_pack missing in master_sheet.' AS error_message
                 FROM matched
                 WHERE format_sku_code IS NOT NULL
                   AND (case_pack IS NULL OR case_pack = 0)
                   AND (
                        NULLIF(TRIM(COALESCE(sku_sap_name, '')), '') IS NOT NULL
                     OR NULLIF(TRIM(COALESCE(sku_sap_code, '')), '') IS NOT NULL
                   )

                UNION ALL

                SELECT raw_row_number,
                       'per_unit_value' AS field_name,
                       'per_unit_value_missing' AS error_type,
                       'per_unit_value missing in master_sheet.' AS error_message
                  FROM matched
                 WHERE format_sku_code IS NOT NULL
                   AND (per_unit_value IS NULL OR per_unit_value = 0)
                   AND (
                        UPPER(COALESCE(uom, '')) IN ('LTR', 'LITRE', 'LITRES', 'ML', 'MLS')
                     OR UPPER(COALESCE(per_unit, '')) LIKE '%%LTR%%'
                     OR UPPER(COALESCE(per_unit, '')) LIKE '%%LITRE%%'
                     OR UPPER(COALESCE(per_unit, '')) LIKE '%%ML%%'
                   )
          ) warnings
         ORDER BY raw_row_number, error_type
        """,
        [upload_id],
    )
    return [
        {
            "row_number": row_number,
            "field_name": field_name,
            "error_type": error_type,
            "error_message": error_message,
            "severity": "warning",
        }
        for row_number, field_name, error_type, error_message in cur.fetchall()
    ]


def _count_issue_rows(issues: list[dict[str, Any]], severity: str) -> int:
    return len(
        {
            issue.get("row_number")
            for issue in issues
            if issue.get("severity") == severity and issue.get("row_number") is not None
        }
    )


def _has_errors(issues: list[dict[str, Any]]) -> bool:
    return any(issue.get("severity") == "error" for issue in issues)


def _transform_appointment(cur, upload_id: int) -> tuple[int, int]:
    cur.execute(
        """
        WITH src AS (
            SELECT *
              FROM staging."appointment data"
             WHERE upload_id = %s
               AND NULLIF(TRIM(appointment_id), '') IS NOT NULL
        ),
        expanded AS (
            SELECT md5(concat_ws('|',
                       LOWER(TRIM(COALESCE(s.appointment_id, ''))),
                       LOWER(TRIM(COALESCE(split_pos, '')))
                   )) AS appointment_line_key,
                   s.appointment_id, s.status, s.appointment_time,
                   s.creation_date, split_pos AS pos, s.destination_fc,
                   s.pro, s.upload_id, s.raw_row_number
              FROM src s
              CROSS JOIN LATERAL (
                  SELECT NULLIF(TRIM(po_value), '') AS split_pos
                    FROM unnest(
                        CASE
                            WHEN NULLIF(TRIM(COALESCE(s.pos, '')), '') IS NULL
                                THEN ARRAY[NULL::text]
                            ELSE regexp_split_to_array(s.pos, '\\s*[,;]\\s*')
                        END
                    ) AS parts(po_value)
                   WHERE NULLIF(TRIM(COALESCE(s.pos, '')), '') IS NULL
                      OR NULLIF(TRIM(po_value), '') IS NOT NULL
              ) po_parts
        ),
        deduped AS (
            SELECT DISTINCT ON (appointment_line_key)
                   appointment_line_key, appointment_id, status, appointment_time,
                   creation_date, pos, destination_fc, pro, upload_id
              FROM expanded
             ORDER BY appointment_line_key, raw_row_number DESC
        ),
        current_appointments AS (
            SELECT DISTINCT LOWER(TRIM(appointment_id)) AS appointment_id_key
              FROM deduped
        ),
        stale_deleted AS (
            DELETE FROM reporting."appointment" existing
             USING current_appointments current_ids
             WHERE LOWER(TRIM(existing.appointment_id)) = current_ids.appointment_id_key
               AND NOT EXISTS (
                   SELECT 1
                     FROM deduped d
                    WHERE d.appointment_line_key = existing.appointment_line_key
               )
             RETURNING 1
        )
        INSERT INTO reporting."appointment" (
            appointment_line_key, appointment_id, status, appointment_time, creation_date, pos,
            destination_fc, pro, month, year, upload_id, updated_at
        )
        SELECT appointment_line_key, appointment_id, status, appointment_time, creation_date, pos,
               destination_fc, pro,
               UPPER(to_char(appointment_time, 'FMMonth')),
               EXTRACT(YEAR FROM appointment_time)::int,
               upload_id, now()
          FROM deduped
        ON CONFLICT (appointment_line_key) DO UPDATE SET
            appointment_id = EXCLUDED.appointment_id,
            status = EXCLUDED.status,
            appointment_time = EXCLUDED.appointment_time,
            creation_date = EXCLUDED.creation_date,
            pos = EXCLUDED.pos,
            destination_fc = EXCLUDED.destination_fc,
            pro = EXCLUDED.pro,
            month = EXCLUDED.month,
            year = EXCLUDED.year,
            upload_id = EXCLUDED.upload_id,
            updated_at = now()
        RETURNING (xmax::text = '0') AS inserted
        """,
        [upload_id],
    )
    flags = [bool(row[0]) for row in cur.fetchall()]
    return flags.count(True), flags.count(False)


def _transform_amazon_po(cur, upload_id: int) -> tuple[int, int]:
    cur.execute(
        """
        WITH src AS (
            SELECT *,
                   md5(concat_ws('|',
                       LOWER(TRIM(COALESCE(po_number, ''))),
                       LOWER(TRIM(COALESCE(asin, '')))
                   )) AS source_line_key
              FROM staging."amazon data"
             WHERE upload_id = %s
        ),
        deduped AS (
            SELECT DISTINCT ON (source_line_key) *
              FROM src
             WHERE NULLIF(TRIM(po_number), '') IS NOT NULL
               AND NULLIF(TRIM(asin), '') IS NOT NULL
               AND NULLIF(TRIM(ship_to_location), '') IS NOT NULL
             ORDER BY source_line_key, raw_row_number DESC
        ),
        enriched AS (
            SELECT s.*,
                   pm.product_name AS master_product_name,
                   pm.item, pm.sku_sap_name AS sap_sku_name,
                   pm.sku_sap_code AS sap_sku_code, pm.category,
                   pm.sub_category, COALESCE(pm.case_pack, NULLIF(s.case_size, 0)) AS case_pack,
                   pm.per_unit_value, pm.per_unit, pm.item_head,
                   pm.tax_rate, pm.brand, pm.category_head, pm.uom,
                   margin.margin_percent AS asin_margin_percent,
                   fc_channel.channel AS core_fresh_now_channel,
                   fc.city, fc.state
              FROM deduped s
              LEFT JOIN LATERAL (
                  SELECT format_sku_code, product_name, item, sku_sap_name,
                         sku_sap_code, category, sub_category, case_pack,
                         COALESCE(
                             NULLIF(per_unit_value::numeric, 0),
                             CASE
                                 WHEN NULLIF(TRIM(per_unit::text), '') IS NULL THEN NULL
                                 WHEN substring(per_unit::text from '([0-9]+(?:\.[0-9]+)?)') IS NULL THEN NULL
                                 WHEN UPPER(COALESCE(uom, '')) IN ('ML', 'MLS')
                                      OR UPPER(per_unit::text) LIKE '%%ML%%'
                                      THEN substring(per_unit::text from '([0-9]+(?:\.[0-9]+)?)')::numeric / 1000
                                 WHEN UPPER(COALESCE(uom, '')) IN ('LTR', 'LITRE', 'LITRES')
                                      OR UPPER(per_unit::text) LIKE '%%LTR%%'
                                      OR UPPER(per_unit::text) LIKE '%%LITRE%%'
                                      THEN substring(per_unit::text from '([0-9]+(?:\.[0-9]+)?)')::numeric
                                 ELSE NULL
                             END
                         ) AS per_unit_value,
                         per_unit,
                         tax_rate::numeric AS tax_rate,
                         item_head, brand, category_head, uom, format
                    FROM public.master_sheet pm
                   WHERE (
                          (NULLIF(s.asin, '') IS NOT NULL AND UPPER(TRIM(pm.format_sku_code::text)) = UPPER(TRIM(s.asin::text)))
                       OR (NULLIF(s.external_id, '') IS NOT NULL AND UPPER(TRIM(pm.format_sku_code::text)) = UPPER(TRIM(s.external_id::text)))
                       OR (NULLIF(s.merchant_sku, '') IS NOT NULL AND UPPER(TRIM(pm.item::text)) = UPPER(TRIM(s.merchant_sku::text)))
                       OR (NULLIF(s.product_name, '') IS NOT NULL AND LOWER(TRIM(pm.product_name::text)) = LOWER(TRIM(s.product_name::text)))
                     )
                   ORDER BY CASE
                       WHEN UPPER(COALESCE(pm.format, '')) = 'AMAZON'
                            AND NULLIF(s.asin, '') IS NOT NULL
                            AND UPPER(TRIM(pm.format_sku_code::text)) = UPPER(TRIM(s.asin::text)) THEN 1
                       WHEN UPPER(COALESCE(pm.format, '')) = 'AMAZON'
                            AND NULLIF(s.external_id, '') IS NOT NULL
                            AND UPPER(TRIM(pm.format_sku_code::text)) = UPPER(TRIM(s.external_id::text)) THEN 2
                       WHEN UPPER(COALESCE(pm.format, '')) = 'AMAZON'
                            AND NULLIF(s.merchant_sku, '') IS NOT NULL
                            AND UPPER(TRIM(pm.item::text)) = UPPER(TRIM(s.merchant_sku::text)) THEN 3
                       WHEN UPPER(COALESCE(pm.format, '')) = 'AMAZON'
                            AND NULLIF(s.product_name, '') IS NOT NULL
                            AND LOWER(TRIM(pm.product_name::text)) = LOWER(TRIM(s.product_name::text)) THEN 4
                       ELSE 5
                   END
                   LIMIT 1
              ) pm ON true
              LEFT JOIN LATERAL (
                  SELECT margin_percent::numeric AS margin_percent
                    FROM public.amazon_asin_margin margin
                   WHERE NULLIF(s.asin, '') IS NOT NULL
                     AND UPPER(TRIM(margin.asin::text)) = UPPER(TRIM(s.asin::text))
                   ORDER BY id
                   LIMIT 1
              ) margin ON true
              LEFT JOIN public.fc_city_state_channel_master fc_channel
                ON NULLIF(s.ship_to_location, '') IS NOT NULL
               AND UPPER(TRIM(fc_channel.fc::text)) = UPPER(TRIM(s.ship_to_location::text))
              LEFT JOIN master.fc_master fc
                ON fc.is_active = true AND fc.fc_code = s.ship_to_location
        ),
        calculated AS (
            SELECT e.*,
                   COALESCE(e.cancellation_deadline, e.window_end, e.expected_date) AS expiry_calc,
                   CASE
                       WHEN e.vendor_code = '0M7KK' THEN 'RK WORLD'
                       ELSE e.vendor_code
                   END AS vendor_calc,
                   CASE
                       WHEN e.per_unit_value IS NULL THEN NULL
                       WHEN e.uom IS NULL OR UPPER(e.uom) = 'LTR'
                           THEN trim(to_char(e.per_unit_value, 'FM999999990.999')) || ' LTR'
                       ELSE trim(to_char(e.per_unit_value, 'FM999999990.999')) || ' ' || UPPER(e.uom)
                   END AS per_ltr_unit_calc,
                   e.per_unit_value AS per_liter_calc
              FROM enriched e
        ),
        statused AS (
            SELECT c.*,
                   GREATEST((c.expiry_calc - CURRENT_DATE)::int, 0) AS days_to_expiry_calc,
                   CASE
                       WHEN TRIM(COALESCE(c.availability, '')) = 'AC - Accepted: In stock'
                            AND COALESCE(c.received_quantity, 0) = 0
                            AND COALESCE(c.accepted_quantity, 0) = 0
                            AND COALESCE(c.cancelled_quantity, 0) = 0
                            AND c.expiry_calc IS NOT NULL
                            AND c.expiry_calc < CURRENT_DATE THEN 'EXPIRED'
                       WHEN TRIM(COALESCE(c.availability, '')) = 'AC - Accepted: In stock'
                            AND TRIM(COALESCE(c.status, '')) = 'Confirmed'
                            AND COALESCE(c.accepted_quantity, 0) > 0
                            AND COALESCE(c.received_quantity, 0) = 0
                            AND COALESCE(c.cancelled_quantity, 0) = 0
                            AND c.expiry_calc IS NOT NULL
                            AND c.expiry_calc < CURRENT_DATE THEN 'EXPIRED'
                       WHEN TRIM(COALESCE(c.status, '')) = 'Confirmed'
                            AND TRIM(COALESCE(c.availability, '')) = 'OS - Cancelled: Out of stock'
                            AND COALESCE(c.accepted_quantity, 0) = 0
                            AND COALESCE(c.received_quantity, 0) = 0
                            AND COALESCE(c.cancelled_quantity, 0) = 0 THEN 'MOV'
                       WHEN TRIM(COALESCE(c.status, '')) = 'Closed'
                            AND TRIM(COALESCE(c.availability, '')) = 'OS - Cancelled: Out of stock'
                            AND COALESCE(c.accepted_quantity, 0) = 0
                            AND COALESCE(c.received_quantity, 0) = 0
                            AND COALESCE(c.cancelled_quantity, 0) = 0 THEN 'CANCELLED'
                       WHEN TRIM(COALESCE(c.availability, '')) = 'AC - Accepted: In stock'
                            AND COALESCE(c.accepted_quantity, 0) > 0
                            AND COALESCE(c.received_quantity, 0) > 0 THEN 'COMPLETED'
                       WHEN TRIM(COALESCE(c.availability, '')) = 'AC - Accepted: In stock'
                            AND TRIM(COALESCE(c.status, '')) = 'Closed'
                            AND COALESCE(c.received_quantity, 0) > 0 THEN 'COMPLETED'
                       WHEN TRIM(COALESCE(c.availability, '')) = 'AC - Accepted: In stock'
                            AND COALESCE(c.cancelled_quantity, 0) > 0
                            AND COALESCE(c.received_quantity, 0) = 0 THEN 'CANCELLED'
                       WHEN TRIM(COALESCE(c.availability, '')) = 'AC - Accepted: In stock'
                            AND TRIM(COALESCE(c.status, '')) = 'Unconfirmed'
                            AND COALESCE(c.accepted_quantity, 0) = 0
                            AND COALESCE(c.received_quantity, 0) = 0
                            AND COALESCE(c.cancelled_quantity, 0) = 0 THEN 'PENDING'
                       WHEN TRIM(COALESCE(c.availability, '')) = 'AC - Accepted: In stock'
                            AND TRIM(COALESCE(c.status, '')) = 'Confirmed'
                            AND COALESCE(c.accepted_quantity, 0) > 0
                            AND COALESCE(c.received_quantity, 0) = 0
                            AND COALESCE(c.cancelled_quantity, 0) = 0 THEN 'PENDING'
                       WHEN TRIM(COALESCE(c.availability, '')) = 'AC - Accepted: In stock'
                            AND TRIM(COALESCE(c.status, '')) = 'Confirmed'
                            AND COALESCE(c.requested_quantity, 0) > 0
                            AND COALESCE(c.accepted_quantity, 0) = 0
                            AND COALESCE(c.received_quantity, 0) = 0
                            AND COALESCE(c.cancelled_quantity, 0) = 0 THEN 'PENDING'
                       ELSE ''
                   END AS po_status_calc
              FROM calculated c
        )
        INSERT INTO reporting."Amazon PO" (
            source_line_key, po_number, order_date, expiry_date, status,
            availability_status, external_id, asin, merchant_sku, sku_code,
            sku_name, requested_qty, accepted_qty, received_qty, cancelled_qty,
            fulfillment_center, cost_price, total_requested_cost, total_accepted_cost,
            total_received_cost, total_cancelled_cost, vendor, days_to_expiry,
            po_window, po_status, item_status, item, sap_sku_name,
            sap_sku_code, category, sub_category, case_pack, requested_boxes,
            accepted_boxes, per_ltr_unit, per_liter, total_order_liters,
            total_accepted_liters, total_delivered_liters,
            total_order_amt_exclusive, total_deliver_amt_exclusive, po_month,
            year, item_head, city, state, distributor_margin,
            tax, brand, category_head, core_fresh_now, order_ltrs_cl,
            missed_ltrs, filled_ltrs, order_unit_cl, missed_unit, filled_units,
            fill_rate, miss_rate, helper, upload_id, updated_at
        )
        SELECT
            source_line_key,
            po_number,
            order_date,
            expiry_calc AS expiry_date,
            status,
            availability AS availability_status,
            external_id,
            asin,
            merchant_sku,
            asin AS sku_code,
            COALESCE(NULLIF(product_name, ''), master_product_name) AS sku_name,
            requested_quantity,
            accepted_quantity,
            received_quantity,
            cancelled_quantity,
            ship_to_location AS fulfillment_center,
            cost AS cost_price,
            total_requested_cost,
            total_accepted_cost,
            total_received_cost,
            total_cancelled_cost,
            vendor_calc AS vendor,
            days_to_expiry_calc AS days_to_expiry,
            (expiry_calc - order_date)::int AS po_window,
            po_status_calc AS po_status,
            CASE
                WHEN po_status_calc = 'COMPLETED' AND COALESCE(received_quantity, 0) >= COALESCE(requested_quantity, 0)
                    THEN 'FULL SUPPLIED'
                WHEN po_status_calc = 'COMPLETED' AND COALESCE(received_quantity, 0) < COALESCE(requested_quantity, 0)
                    THEN 'SHORT SUPPLIED'
                ELSE NULL
            END AS item_status,
            item,
            sap_sku_name,
            sap_sku_code,
            category,
            sub_category,
            case_pack,
            requested_quantity / NULLIF(case_pack, 0) AS requested_boxes,
            accepted_quantity / NULLIF(case_pack, 0) AS accepted_boxes,
            per_unit AS per_ltr_unit,
            per_liter_calc AS per_liter,
            requested_quantity * COALESCE(per_liter_calc, 0) AS total_order_liters,
            accepted_quantity * COALESCE(per_liter_calc, 0) AS total_accepted_liters,
            received_quantity * COALESCE(per_liter_calc, 0) AS total_delivered_liters,
            total_requested_cost AS total_order_amt_exclusive,
            total_received_cost AS total_deliver_amt_exclusive,
            EXTRACT(MONTH FROM order_date)::int AS po_month,
            EXTRACT(YEAR FROM order_date)::int AS year,
            item_head,
            city,
            state,
            asin_margin_percent AS distributor_margin,
            tax_rate AS tax,
            brand,
            category_head,
            core_fresh_now_channel AS core_fresh_now,
            requested_quantity * COALESCE(per_liter_calc, 0) AS order_ltrs_cl,
            CASE
                WHEN po_status_calc IN ('MOV', 'CANCELLED') THEN NULL
                WHEN po_status_calc IN ('COMPLETED', 'EXPIRED') THEN (requested_quantity - received_quantity) * COALESCE(per_liter_calc, 0)
                ELSE 0
            END AS missed_ltrs,
            received_quantity * COALESCE(per_liter_calc, 0) AS filled_ltrs,
            requested_quantity AS order_unit_cl,
            CASE
                WHEN po_status_calc IN ('MOV', 'CANCELLED') THEN NULL
                WHEN po_status_calc IN ('COMPLETED', 'EXPIRED') THEN requested_quantity - received_quantity
                ELSE 0
            END AS missed_unit,
            received_quantity AS filled_units,
            received_quantity / NULLIF(requested_quantity, 0) AS fill_rate,
            CASE
                WHEN po_status_calc = 'CANCELLED' THEN NULL
                WHEN po_status_calc IN ('PENDING', 'MOV') THEN 0
                ELSE 1 - (received_quantity / NULLIF(requested_quantity, 0))
            END AS miss_rate,
            CASE
                WHEN LOWER(COALESCE(status, '')) = 'confirmed'
                     AND po_status_calc IN ('PENDING', 'MOV')
                     AND days_to_expiry_calc BETWEEN 1 AND 18 THEN 'INCLUDE'
                ELSE 'EXCLUDE'
            END AS helper,
            upload_id,
            now()
          FROM statused
        ON CONFLICT (source_line_key) DO UPDATE SET
            po_number = EXCLUDED.po_number,
            order_date = EXCLUDED.order_date,
            expiry_date = EXCLUDED.expiry_date,
            status = EXCLUDED.status,
            availability_status = EXCLUDED.availability_status,
            external_id = EXCLUDED.external_id,
            asin = EXCLUDED.asin,
            merchant_sku = EXCLUDED.merchant_sku,
            sku_code = EXCLUDED.sku_code,
            sku_name = EXCLUDED.sku_name,
            requested_qty = EXCLUDED.requested_qty,
            accepted_qty = EXCLUDED.accepted_qty,
            received_qty = EXCLUDED.received_qty,
            cancelled_qty = EXCLUDED.cancelled_qty,
            fulfillment_center = EXCLUDED.fulfillment_center,
            cost_price = EXCLUDED.cost_price,
            total_requested_cost = EXCLUDED.total_requested_cost,
            total_accepted_cost = EXCLUDED.total_accepted_cost,
            total_received_cost = EXCLUDED.total_received_cost,
            total_cancelled_cost = EXCLUDED.total_cancelled_cost,
            vendor = EXCLUDED.vendor,
            days_to_expiry = EXCLUDED.days_to_expiry,
            po_window = EXCLUDED.po_window,
            po_status = EXCLUDED.po_status,
            item_status = EXCLUDED.item_status,
            item = EXCLUDED.item,
            sap_sku_name = EXCLUDED.sap_sku_name,
            sap_sku_code = EXCLUDED.sap_sku_code,
            category = EXCLUDED.category,
            sub_category = EXCLUDED.sub_category,
            case_pack = EXCLUDED.case_pack,
            requested_boxes = EXCLUDED.requested_boxes,
            accepted_boxes = EXCLUDED.accepted_boxes,
            per_ltr_unit = EXCLUDED.per_ltr_unit,
            per_liter = EXCLUDED.per_liter,
            total_order_liters = EXCLUDED.total_order_liters,
            total_accepted_liters = EXCLUDED.total_accepted_liters,
            total_delivered_liters = EXCLUDED.total_delivered_liters,
            total_order_amt_exclusive = EXCLUDED.total_order_amt_exclusive,
            total_deliver_amt_exclusive = EXCLUDED.total_deliver_amt_exclusive,
            po_month = EXCLUDED.po_month,
            year = EXCLUDED.year,
            item_head = EXCLUDED.item_head,
            city = EXCLUDED.city,
            state = EXCLUDED.state,
            distributor_margin = EXCLUDED.distributor_margin,
            tax = EXCLUDED.tax,
            brand = EXCLUDED.brand,
            category_head = EXCLUDED.category_head,
            core_fresh_now = EXCLUDED.core_fresh_now,
            order_ltrs_cl = EXCLUDED.order_ltrs_cl,
            missed_ltrs = EXCLUDED.missed_ltrs,
            filled_ltrs = EXCLUDED.filled_ltrs,
            order_unit_cl = EXCLUDED.order_unit_cl,
            missed_unit = EXCLUDED.missed_unit,
            filled_units = EXCLUDED.filled_units,
            fill_rate = EXCLUDED.fill_rate,
            miss_rate = EXCLUDED.miss_rate,
            helper = EXCLUDED.helper,
            upload_id = EXCLUDED.upload_id,
            updated_at = now()
        RETURNING (xmax::text = '0') AS inserted
        """,
        [upload_id],
    )
    flags = [bool(row[0]) for row in cur.fetchall()]
    return flags.count(True), flags.count(False)


def _run_transform(cur, *, config: ReportConfig, upload_id: int) -> tuple[int, int]:
    if config.report_type == "APPOINTMENT":
        return _transform_appointment(cur, upload_id)
    if config.report_type == "AMAZON_PO":
        return _transform_amazon_po(cur, upload_id)
    raise ValueError("Unsupported report_type.")


def _response_payload(
    *,
    upload_id: int,
    config: ReportConfig,
    status_value: str,
    rows_received: int,
    rows_inserted_staging: int,
    inserted: int,
    updated: int,
    issues: list[dict[str, Any]],
) -> dict[str, Any]:
    errors = [issue for issue in issues if issue.get("severity") == "error"]
    warnings = [issue for issue in issues if issue.get("severity") == "warning"]
    return {
        "upload_id": upload_id,
        "report_type": config.report_type,
        "main_table_name": config.main_table_name,
        "raw_file_name": config.raw_file_name,
        "status": status_value,
        "rows_received": rows_received,
        "rows_inserted_staging": rows_inserted_staging,
        "rows_inserted_final": inserted,
        "rows_updated_final": updated,
        "error_count": len(errors),
        "warning_count": len(warnings),
        "errors": errors[:200],
        "warnings": warnings[:200],
    }


def process_upload(request) -> tuple[dict[str, Any], int]:
    report_type = str(request.data.get("report_type") or "").strip().upper()
    config = REPORTS.get(report_type)
    if not config:
        return {"detail": "Unsupported report_type."}, status.HTTP_400_BAD_REQUEST

    upload = request.FILES.get("file")
    pasted_data = str(request.data.get("pasted_data") or "").strip()
    upload_source = "file" if upload else "paste"

    if upload:
        original_name = Path(upload.name or "upload").name
        extension = Path(original_name).suffix.lower()
        if extension not in ALLOWED_EXTENSIONS:
            return {"detail": "Only CSV and XLSX uploads are supported."}, status.HTTP_400_BAD_REQUEST
        content = upload.read()
    elif pasted_data:
        original_name = (
            str(request.data.get("original_file_name") or f"{config.raw_file_name}-pasted.csv")
            .strip()
            .replace("/", "-")
            .replace("\\", "-")
        )
        extension = ".csv"
        content = (pasted_data + "\n").encode("utf-8")
    else:
        return {"detail": "Paste data or choose a CSV/XLSX file."}, status.HTTP_400_BAD_REQUEST

    if len(content) > MAX_UPLOAD_SIZE:
        return {"detail": "Upload is too large. Maximum size is 20 MB."}, status.HTTP_400_BAD_REQUEST

    uploaded_by = str(
        request.data.get("uploaded_by")
        or getattr(request.user, "email", "")
        or getattr(request.user, "username", "")
        or "unknown"
    ).strip()
    file_hash = hashlib.sha256(content).hexdigest()
    stored_file_path = _store_file(content, original_name)
    base_metadata = {
        "report_type": config.report_type,
        "upload_source": upload_source,
    }

    with transaction.atomic():
        with connection.cursor() as cur:
            cur.execute(
                """
                SELECT upload_id, status
                  FROM raw.upload_file
                 WHERE file_hash = %s
                   AND main_table_name = %s
                   AND raw_file_name = %s
                   AND status IN (
                       'completed', 'partially_successful', 'staged',
                       'uploaded', 'validating'
                   )
                 ORDER BY uploaded_at DESC
                 LIMIT 1
                """,
                [
                    file_hash,
                    config.main_table_name,
                    config.raw_file_name,
                ],
            )
            duplicate = cur.fetchone()
            if duplicate:
                base_metadata = {
                    **base_metadata,
                    "duplicate_of": duplicate[0],
                    "duplicate_processed": True,
                }

            upload_id = _insert_upload_file(
                cur,
                config=config,
                original_file_name=original_name,
                stored_file_path=stored_file_path,
                file_hash=file_hash,
                file_extension=extension,
                uploaded_by=uploaded_by,
                status_value="uploaded",
                metadata=base_metadata,
            )
            _update_upload_file(
                cur,
                upload_id=upload_id,
                status_value="validating",
                row_count=0,
                error_count=0,
                warning_count=0,
            )

            try:
                rows, issues, rows_received = parse_uploaded_file(
                    config=config,
                    content=content,
                    extension=extension,
                )
            except ValueError as exc:
                issues = [
                    {
                        "row_number": None,
                        "field_name": None,
                        "error_type": "parse_failed",
                        "error_message": str(exc),
                        "severity": "error",
                    }
                ]
                rows = []
                rows_received = 0

            if rows_received == 0 and not any(
                issue.get("error_type") in {"empty_file", "parse_failed"} for issue in issues
            ):
                issues.append(
                    {
                        "row_number": None,
                        "field_name": None,
                        "error_type": "no_data_rows",
                        "error_message": (
                            "No data rows were found. Paste the header row first, "
                            "then at least one data row copied from Excel or Sheets."
                        ),
                        "severity": "error",
                    }
                )
            elif rows_received > 0 and not rows:
                issues.append(
                    {
                        "row_number": None,
                        "field_name": None,
                        "error_type": "no_valid_rows",
                        "error_message": (
                            "No usable rows were parsed. Check that the first pasted row "
                            "contains the column headers for the selected report type."
                        ),
                        "severity": "error",
                    }
                )

            rows_inserted_staging = _insert_staging_rows(
                cur,
                upload_id=upload_id,
                config=config,
                rows=rows,
            )
            if rows_inserted_staging != len(rows):
                issues.append(
                    {
                        "row_number": None,
                        "field_name": None,
                        "error_type": "staging_row_count_mismatch",
                        "error_message": "Parser row count did not match staging insert count.",
                        "severity": "error",
                    }
                )
            issues.extend(
                _add_business_warnings(
                    cur,
                    config=config,
                    upload_id=upload_id,
                )
            )

            _insert_validation_issues(cur, upload_id=upload_id, config=config, issues=issues)
            error_count = sum(1 for issue in issues if issue.get("severity") == "error")
            warning_count = sum(1 for issue in issues if issue.get("severity") == "warning")

            inserted = 0
            updated = 0
            if _has_errors(issues):
                status_value = "failed"
            else:
                _update_upload_file(
                    cur,
                    upload_id=upload_id,
                    status_value="staged",
                    row_count=rows_received,
                    error_count=error_count,
                    warning_count=warning_count,
                )
                try:
                    inserted, updated = _run_transform(
                        cur,
                        config=config,
                        upload_id=upload_id,
                    )
                    status_value = "partially_successful" if warning_count else "completed"
                except Exception:
                    issues.append(
                        {
                            "row_number": None,
                            "field_name": None,
                            "error_type": "transform_failed",
                            "error_message": "Could not transform staged rows into the reporting table.",
                            "severity": "error",
                        }
                    )
                    _insert_validation_issues(
                        cur,
                        upload_id=upload_id,
                        config=config,
                        issues=[issues[-1]],
                    )
                    error_count += 1
                    status_value = "failed"

            valid_rows = max(rows_received - _count_issue_rows(issues, "error"), 0)
            _upsert_summary(
                cur,
                upload_id=upload_id,
                config=config,
                total_rows=rows_received,
                valid_rows=valid_rows,
                error_rows=_count_issue_rows(issues, "error"),
                warning_rows=_count_issue_rows(issues, "warning"),
                inserted=inserted,
                updated=updated,
                status_value=status_value,
            )
            _update_upload_file(
                cur,
                upload_id=upload_id,
                status_value=status_value,
                row_count=rows_received,
                error_count=error_count,
                warning_count=warning_count,
                metadata={
                    "rows_inserted_staging": rows_inserted_staging,
                    "rows_inserted_final": inserted,
                    "rows_updated_final": updated,
                },
            )

            http_status = status.HTTP_400_BAD_REQUEST if status_value == "failed" else status.HTTP_200_OK
            return (
                _response_payload(
                    upload_id=upload_id,
                    config=config,
                    status_value=status_value,
                    rows_received=rows_received,
                    rows_inserted_staging=rows_inserted_staging,
                    inserted=inserted,
                    updated=updated,
                    issues=issues,
                ),
                http_status,
            )


def _jsonable(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _json_object(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            return {}
        return parsed if isinstance(parsed, dict) else {}
    return {}


def _upload_column_check(upload: dict[str, Any], config: ReportConfig) -> dict[str, Any]:
    path = upload.get("stored_file_path")
    extension = str(upload.get("file_extension") or Path(path or "").suffix).lower()
    if extension and not extension.startswith("."):
        extension = f".{extension}"
    base = {
        "available": False,
        "required_columns": list(config.required_columns),
        "identifier_columns": ["asin"] if config.report_type == "AMAZON_PO" else [],
        "found_fields": [],
        "missing_required_columns": list(config.required_columns),
        "missing_identifier_columns": [],
        "original_headers": [],
        "message": "",
    }
    if not path or not Path(path).exists():
        return {**base, "message": "Stored upload file is not available for header checking."}
    try:
        content = Path(path).read_bytes()
        table = _read_csv(content) if extension == ".csv" else _read_xlsx(content)
    except Exception as exc:
        return {**base, "message": f"Could not read stored upload headers: {exc}"}

    table = [row for row in table if any(not _is_blank(cell) for cell in row)]
    if not table:
        return {**base, "message": "Stored upload file does not contain header rows."}

    original_headers = ["" if _is_blank(cell) else str(cell).strip() for cell in table[0]]
    aliases = _alias_map(config)
    found_fields = sorted(
        {
            aliases[normalize_header(header)]
            for header in original_headers
            if aliases.get(normalize_header(header))
        }
    )
    missing_required = [field for field in config.required_columns if field not in found_fields]
    identifier_columns = base["identifier_columns"]
    missing_identifier = (
        identifier_columns
        if identifier_columns and not set(identifier_columns).intersection(found_fields)
        else []
    )
    return {
        **base,
        "available": True,
        "found_fields": found_fields,
        "missing_required_columns": missing_required,
        "missing_identifier_columns": missing_identifier,
        "original_headers": original_headers,
        "message": "",
    }


def _upload_issue_groups(cur, upload_id: int) -> list[dict[str, Any]]:
    cur.execute(
        """
        SELECT severity, field_name, error_type, COUNT(*) AS issue_count,
               MIN(row_number) AS first_row, MAX(row_number) AS last_row
          FROM quality.validation_error
         WHERE upload_id = %s
         GROUP BY severity, field_name, error_type
         ORDER BY CASE severity WHEN 'error' THEN 0 ELSE 1 END,
                  issue_count DESC, error_type, field_name
        """,
        [upload_id],
    )
    cols = [desc[0] for desc in cur.description]
    return [{col: _jsonable(value) for col, value in zip(cols, row)} for row in cur.fetchall()]


def _amazon_upload_diagnostics(cur, upload_id: int, summary: dict[str, Any] | None) -> dict[str, Any]:
    cur.execute(
        """
        SELECT COUNT(*) AS raw_rows,
               COUNT(DISTINCT md5(concat_ws('|',
                   LOWER(TRIM(COALESCE(po_number, ''))),
                   LOWER(TRIM(COALESCE(asin, '')))
               ))) FILTER (
                   WHERE NULLIF(TRIM(po_number), '') IS NOT NULL
                     AND NULLIF(TRIM(asin), '') IS NOT NULL
               ) AS unique_final_keys
          FROM staging."amazon data"
         WHERE upload_id = %s
        """,
        [upload_id],
    )
    raw_rows, unique_final_keys = cur.fetchone()
    raw_rows = int(raw_rows or 0)
    unique_final_keys = int(unique_final_keys or 0)

    cur.execute(
        """
        SELECT s.ship_to_location AS fc_code, COUNT(*) AS row_count
          FROM staging."amazon data" s
          LEFT JOIN master.fc_master fc
            ON fc.is_active = true AND fc.fc_code = s.ship_to_location
         WHERE s.upload_id = %s
           AND NULLIF(TRIM(s.ship_to_location), '') IS NOT NULL
           AND fc.fc_id IS NULL
         GROUP BY s.ship_to_location
         ORDER BY row_count DESC, s.ship_to_location
         LIMIT 20
        """,
        [upload_id],
    )
    missing_fcs = _rows_to_dicts(cur)

    cur.execute(
        """
        SELECT s.asin, s.external_id, s.merchant_sku,
               LEFT(COALESCE(s.product_name, ''), 120) AS product_name,
               COUNT(*) AS row_count
          FROM staging."amazon data" s
          LEFT JOIN LATERAL (
              SELECT pm.format_sku_code
                FROM public.master_sheet pm
               WHERE (
                      (NULLIF(s.asin, '') IS NOT NULL AND UPPER(TRIM(pm.format_sku_code::text)) = UPPER(TRIM(s.asin::text)))
                   OR (NULLIF(s.external_id, '') IS NOT NULL AND UPPER(TRIM(pm.format_sku_code::text)) = UPPER(TRIM(s.external_id::text)))
                   OR (NULLIF(s.merchant_sku, '') IS NOT NULL AND UPPER(TRIM(pm.item::text)) = UPPER(TRIM(s.merchant_sku::text)))
                   OR (NULLIF(s.product_name, '') IS NOT NULL AND LOWER(TRIM(pm.product_name::text)) = LOWER(TRIM(s.product_name::text)))
                 )
               ORDER BY CASE
                   WHEN UPPER(COALESCE(pm.format, '')) = 'AMAZON'
                        AND NULLIF(s.asin, '') IS NOT NULL
                        AND UPPER(TRIM(pm.format_sku_code::text)) = UPPER(TRIM(s.asin::text)) THEN 1
                   WHEN UPPER(COALESCE(pm.format, '')) = 'AMAZON'
                        AND NULLIF(s.external_id, '') IS NOT NULL
                        AND UPPER(TRIM(pm.format_sku_code::text)) = UPPER(TRIM(s.external_id::text)) THEN 2
                   WHEN UPPER(COALESCE(pm.format, '')) = 'AMAZON'
                        AND NULLIF(s.merchant_sku, '') IS NOT NULL
                        AND UPPER(TRIM(pm.item::text)) = UPPER(TRIM(s.merchant_sku::text)) THEN 3
                   WHEN UPPER(COALESCE(pm.format, '')) = 'AMAZON'
                        AND NULLIF(s.product_name, '') IS NOT NULL
                        AND LOWER(TRIM(pm.product_name::text)) = LOWER(TRIM(s.product_name::text)) THEN 4
                   ELSE 5
               END
               LIMIT 1
          ) pm ON true
         WHERE s.upload_id = %s
           AND pm.format_sku_code IS NULL
         GROUP BY s.asin, s.external_id, s.merchant_sku, s.product_name
         ORDER BY row_count DESC, s.asin NULLS LAST
         LIMIT 20
        """,
        [upload_id],
    )
    missing_products = _rows_to_dicts(cur)

    inserted = int((summary or {}).get("final_inserted_rows") or 0)
    updated = int((summary or {}).get("final_updated_rows") or 0)
    return {
        "row_difference": {
            "raw_rows": raw_rows,
            "unique_final_keys": unique_final_keys,
            "collapsed_rows": max(raw_rows - unique_final_keys, 0),
            "inserted_rows": inserted,
            "updated_rows": updated,
        },
        "missing_fcs": missing_fcs,
        "missing_products": missing_products,
    }


def _upload_diagnostics(
    cur,
    upload: dict[str, Any],
    summary: dict[str, Any] | None,
) -> dict[str, Any]:
    config = REPORTS.get(upload.get("report_type") or "")
    diagnostics: dict[str, Any] = {
        "issue_groups": _upload_issue_groups(cur, int(upload["upload_id"])),
    }
    if config:
        diagnostics["column_check"] = _upload_column_check(upload, config)
        if config.report_type == "AMAZON_PO":
            diagnostics["amazon_po"] = _amazon_upload_diagnostics(
                cur,
                int(upload["upload_id"]),
                summary,
            )
    return diagnostics


def _rows_to_dicts(cur) -> list[dict[str, Any]]:
    cols = [desc[0] for desc in cur.description]
    return [{col: _jsonable(value) for col, value in zip(cols, row)} for row in cur.fetchall()]


def _page_params(request) -> tuple[int, int, int]:
    try:
        page = max(0, int(request.query_params.get("page", 0)))
        page_size = min(10000, max(1, int(request.query_params.get("page_size", 50))))
    except ValueError:
        page, page_size = 0, 50
    return page, page_size, page * page_size


def _add_ilike(where: list[str], params: list[Any], column_sql: str, value: str | None) -> None:
    if value:
        where.append(f"{column_sql} ILIKE %s")
        params.append(f"%{value[:200]}%")


def _add_date_range(
    where: list[str],
    params: list[Any],
    column_sql: str,
    start: str | None,
    end: str | None,
) -> None:
    if start:
        where.append(f"{column_sql} >= %s")
        params.append(start)
    if end:
        where.append(f"{column_sql} <= %s")
        params.append(end)


def _paginated_select(
    *,
    table_sql: str,
    columns: tuple[str, ...],
    where: list[str],
    params: list[Any],
    order_sql: str,
    page: int,
    page_size: int,
    offset: int,
) -> dict[str, Any]:
    where_sql = f" WHERE {' AND '.join(where)}" if where else ""
    column_sql = ", ".join(f'"{col}"' for col in columns)
    with connection.cursor() as cur:
        cur.execute(f"SELECT COUNT(*) FROM {table_sql}{where_sql}", params)
        total = int(cur.fetchone()[0] or 0)
        cur.execute(
            f"SELECT {column_sql} FROM {table_sql}{where_sql} {order_sql} LIMIT %s OFFSET %s",
            params + [page_size, offset],
        )
        results = _rows_to_dicts(cur)
    return {"results": results, "count": total, "page": page, "page_size": page_size}


def _ensure_amazon_access(user) -> None:
    if not can_access_platform(user, "amazon"):
        raise PermissionDenied("You do not have access to the Amazon platform.")


@api_view(["GET", "POST"])
@permission_classes([require("upload.use")])
def uploads_collection(request):
    if request.method == "POST":
        payload, http_status = process_upload(request)
        return Response(payload, status=http_status)

    report_type = str(request.query_params.get("report_type") or "").strip().upper()
    status_filter = str(request.query_params.get("status") or "").strip()
    date_from = request.query_params.get("date_from")
    date_to = request.query_params.get("date_to")
    page, page_size, offset = _page_params(request)

    where: list[str] = []
    params: list[Any] = []
    if report_type:
        config = REPORTS.get(report_type)
        if not config:
            return Response({"detail": "Unsupported report_type."}, status=400)
        where.append("main_table_name = %s")
        params.append(config.main_table_name)
    else:
        supported_main_tables = [cfg.main_table_name for cfg in REPORTS.values()]
        placeholders = ", ".join(["%s"] * len(supported_main_tables))
        where.append(f"main_table_name IN ({placeholders})")
        params.extend(supported_main_tables)
    if status_filter:
        where.append("status = %s")
        params.append(status_filter)
    _add_date_range(where, params, "uploaded_at", date_from, date_to)
    where_sql = f" WHERE {' AND '.join(where)}" if where else ""

    with connection.cursor() as cur:
        cur.execute(f"SELECT COUNT(*) FROM raw.upload_file{where_sql}", params)
        total = int(cur.fetchone()[0] or 0)
        cur.execute(
            f"""
            SELECT upload_id, main_table_name, raw_file_name, original_file_name,
                   uploaded_by, uploaded_at, status, row_count, error_count,
                   warning_count, metadata
              FROM raw.upload_file
              {where_sql}
             ORDER BY uploaded_at DESC, upload_id DESC
             LIMIT %s OFFSET %s
            """,
            params + [page_size, offset],
        )
        rows = _rows_to_dicts(cur)

    for row in rows:
        metadata = _json_object(row.get("metadata"))
        row["metadata"] = metadata
        row["report_type"] = metadata.get(
            "report_type",
            MAIN_TO_REPORT_TYPE.get(row.get("main_table_name")),
        )
    return Response({"results": rows, "count": total, "page": page, "page_size": page_size})


@api_view(["GET"])
@permission_classes([require("upload.use")])
def upload_detail(request, upload_id: int):
    with connection.cursor() as cur:
        cur.execute(
            """
            SELECT upload_id, main_table_name, raw_file_name, original_file_name,
                   stored_file_path, file_hash, file_extension, uploaded_by,
                   uploaded_at, status, row_count, error_count, warning_count,
                   metadata, created_at, updated_at
              FROM raw.upload_file
             WHERE upload_id = %s
            """,
            [upload_id],
        )
        upload_rows = _rows_to_dicts(cur)
        if not upload_rows:
            return Response({"detail": "Upload not found."}, status=404)

        cur.execute(
            """
            SELECT summary_id, upload_id, main_table_name, raw_file_name,
                   total_rows, valid_rows, error_rows, warning_rows,
                   final_inserted_rows, final_updated_rows, status, created_at
              FROM quality.upload_validation_summary
             WHERE upload_id = %s
            """,
            [upload_id],
        )
        summary_rows = _rows_to_dicts(cur)

        cur.execute(
            """
            SELECT validation_error_id, upload_id, main_table_name, raw_file_name,
                   row_number, field_name, error_type, error_message, severity,
                   created_at
              FROM quality.validation_error
             WHERE upload_id = %s
             ORDER BY CASE severity WHEN 'error' THEN 0 ELSE 1 END,
                      row_number NULLS FIRST, validation_error_id
            """,
            [upload_id],
        )
        errors = _rows_to_dicts(cur)

        upload = upload_rows[0]
        metadata = _json_object(upload.get("metadata"))
        upload["metadata"] = metadata
        upload["report_type"] = metadata.get(
            "report_type",
            MAIN_TO_REPORT_TYPE.get(upload.get("main_table_name")),
        )
        summary = summary_rows[0] if summary_rows else None
        diagnostics = _upload_diagnostics(cur, upload, summary)

    return Response(
        {
            "upload": upload,
            "summary": summary,
            "errors": errors,
            "diagnostics": diagnostics,
        }
    )


AMAZON_PO_REPORT_COLUMNS = (
    "po_number",
    "order_date",
    "expiry_date",
    "status",
    "availability_status",
    "external_id",
    "sku_code",
    "sku_name",
    "requested_qty",
    "accepted_qty",
    "received_qty",
    "cancelled_qty",
    "fulfillment_center",
    "cost_price",
    "total_requested_cost",
    "total_accepted_cost",
    "total_received_cost",
    "total_cancelled_cost",
    "vendor",
    "days_to_expiry",
    "po_window",
    "po_status",
    "item_status",
    "item",
    "sap_sku_name",
    "sap_sku_code",
    "category",
    "sub_category",
    "case_pack",
    "requested_boxes",
    "accepted_boxes",
    "per_ltr_unit",
    "per_liter",
    "total_order_liters",
    "total_accepted_liters",
    "total_delivered_liters",
    "total_order_amt_exclusive",
    "total_deliver_amt_exclusive",
    "po_month",
    "year",
    "item_head",
    "city",
    "state",
    "distributor_margin",
    "tax",
    "brand",
    "category_head",
    "core_fresh_now",
    "order_ltrs_cl",
    "missed_ltrs",
    "filled_ltrs",
    "order_unit_cl",
    "missed_unit",
    "filled_units",
    "fill_rate",
    "miss_rate",
    "helper",
)

APPOINTMENT_REPORT_COLUMNS = (
    "appointment_id",
    "status",
    "appointment_time",
    "creation_date",
    "pos",
    "destination_fc",
    "pro",
    "month",
    "year",
)

@api_view(["GET"])
@permission_classes([require("platform.po.view")])
def amazon_po_report(request):
    _ensure_amazon_access(request.user)
    page, page_size, offset = _page_params(request)
    q = request.query_params
    where: list[str] = []
    params: list[Any] = []
    _add_ilike(where, params, "po_number", q.get("po_number"))
    _add_ilike(where, params, "asin", q.get("asin"))
    _add_ilike(where, params, "fulfillment_center", q.get("fulfillment_center"))
    _add_ilike(where, params, "vendor", q.get("vendor"))
    _add_ilike(where, params, "category", q.get("category"))
    _add_ilike(where, params, "item_head", q.get("item_head"))
    _add_ilike(where, params, "state", q.get("state"))
    _add_ilike(where, params, "city", q.get("city"))
    po_status_val = str(q.get("po_status") or "").strip().upper()
    if po_status_val:
        where.append("po_status = %s")
        params.append(po_status_val)
    helper_val = str(q.get("helper") or "").strip().upper()
    if helper_val:
        where.append("helper = %s")
        params.append(helper_val)
    search = q.get("search")
    if search:
        where.append(
            "(po_number ILIKE %s OR external_id ILIKE %s OR asin ILIKE %s OR merchant_sku ILIKE %s OR sku_name ILIKE %s)"
        )
        params.extend([f"%{search[:200]}%"] * 5)
    _add_date_range(where, params, "order_date", q.get("order_date_from"), q.get("order_date_to"))
    sort_by = str(q.get("sort_by") or "").strip().lower()
    order_sql = (
        "ORDER BY expiry_date ASC NULLS LAST, po_number ASC"
        if sort_by == "expiry_date"
        else "ORDER BY order_date DESC NULLS LAST, po_number ASC"
    )
    return Response(
        _paginated_select(
            table_sql='reporting."Amazon PO"',
            columns=AMAZON_PO_REPORT_COLUMNS,
            where=where,
            params=params,
            order_sql=order_sql,
            page=page,
            page_size=page_size,
            offset=offset,
        )
    )


@api_view(["GET"])
@permission_classes([require("platform.po.view")])
def amazon_po_filter_options(request):
    _ensure_amazon_access(request.user)
    with connection.cursor() as cur:
        cur.execute(
            """
            WITH options(value) AS (
                SELECT asin::text
                  FROM reporting."Amazon PO"
                 WHERE asin IS NOT NULL AND TRIM(asin::text) != ''
                UNION
                SELECT asin::text
                  FROM staging."amazon data"
                 WHERE asin IS NOT NULL AND TRIM(asin::text) != ''
                UNION
                SELECT format_sku_code::text
                  FROM public.master_sheet
                 WHERE UPPER(COALESCE(format, '')) = 'AMAZON'
                   AND format_sku_code IS NOT NULL
                   AND TRIM(format_sku_code::text) != ''
            )
            SELECT DISTINCT value
              FROM options
             WHERE value IS NOT NULL AND TRIM(value) != ''
             ORDER BY value ASC
             LIMIT 5000
            """
        )
        asins = [row[0] for row in cur.fetchall()]

        cur.execute(
            """
            WITH options(value) AS (
                SELECT fulfillment_center::text
                  FROM reporting."Amazon PO"
                 WHERE fulfillment_center IS NOT NULL AND TRIM(fulfillment_center::text) != ''
                UNION
                SELECT ship_to_location::text
                  FROM staging."amazon data"
                 WHERE ship_to_location IS NOT NULL AND TRIM(ship_to_location::text) != ''
                UNION
                SELECT fc::text
                  FROM public.fc_city_state_channel_master
                 WHERE fc IS NOT NULL AND TRIM(fc::text) != ''
            )
            SELECT DISTINCT value
              FROM options
             WHERE value IS NOT NULL AND TRIM(value) != ''
             ORDER BY value ASC
             LIMIT 1000
            """
        )
        fulfillment_centers = [row[0] for row in cur.fetchall()]

        cur.execute(
            """
            WITH options(value) AS (
                SELECT po_status::text
                  FROM reporting."Amazon PO"
                 WHERE po_status IS NOT NULL AND TRIM(po_status::text) != ''
            )
            SELECT DISTINCT value
              FROM options
             WHERE value IS NOT NULL AND TRIM(value) != ''
             ORDER BY value ASC
             LIMIT 500
            """
        )
        po_statuses = [row[0] for row in cur.fetchall()]

    return Response(
        {
            "asins": asins,
            "fulfillment_centers": fulfillment_centers,
            "po_statuses": po_statuses,
            "item_heads": ["PREMIUM", "COMMODITY", "OTHER"],
        }
    )


@api_view(["GET"])
@permission_classes([require("platform.po.view")])
def appointment_report(request):
    _ensure_amazon_access(request.user)
    page, page_size, offset = _page_params(request)
    q = request.query_params
    where: list[str] = []
    params: list[Any] = []
    _add_ilike(where, params, "appointment_id", q.get("appointment_id"))
    _add_ilike(where, params, "status", q.get("status"))
    _add_ilike(where, params, "destination_fc", q.get("destination_fc"))
    _add_ilike(where, params, "pro", q.get("pro"))
    status_exact = str(q.get("status_exact") or "").strip()
    if status_exact:
        where.append("LOWER(status) = %s")
        params.append(status_exact.lower())
    search = q.get("search")
    if search:
        where.append(
            "(appointment_id ILIKE %s OR destination_fc ILIKE %s OR pos ILIKE %s OR pro ILIKE %s)"
        )
        params.extend([f"%{search[:200]}%"] * 4)
    _add_date_range(
        where,
        params,
        "appointment_time",
        q.get("appointment_time_from"),
        q.get("appointment_time_to"),
    )
    return Response(
        _paginated_select(
            table_sql='reporting."appointment"',
            columns=APPOINTMENT_REPORT_COLUMNS,
            where=where,
            params=params,
            order_sql="ORDER BY appointment_time DESC NULLS LAST, appointment_id ASC",
            page=page,
            page_size=page_size,
            offset=offset,
        )
    )


@api_view(["GET"])
@permission_classes([require("platform.po.view")])
def product_master_lookup(request):
    _ensure_amazon_access(request.user)
    page, page_size, offset = _page_params(request)
    search = request.query_params.get("search")
    where = []
    params: list[Any] = []
    if search:
        where.append(
            "(format_sku_code ILIKE %s OR product_name ILIKE %s OR item ILIKE %s OR sku_sap_code ILIKE %s OR sku_sap_name ILIKE %s)"
        )
        params.extend([f"%{search[:200]}%"] * 5)
    return Response(
        _paginated_select(
            table_sql="public.master_sheet",
            columns=(
                "format",
                "format_sku_code",
                "product_name",
                "item",
                "sku_sap_code",
                "sku_sap_name",
                "case_pack",
                "per_unit",
                "per_unit_value",
                "tax_rate",
                "uom",
                "item_head",
            ),
            where=where,
            params=params,
            order_sql="ORDER BY CASE WHEN UPPER(COALESCE(format, '')) = 'AMAZON' THEN 0 ELSE 1 END, product_name ASC NULLS LAST, format_sku_code ASC NULLS LAST",
            page=page,
            page_size=page_size,
            offset=offset,
        )
    )


@api_view(["GET"])
@permission_classes([require("platform.po.view")])
def fc_master_lookup(request):
    _ensure_amazon_access(request.user)
    page, page_size, offset = _page_params(request)
    search = request.query_params.get("search")
    where = ["is_active = true"]
    params: list[Any] = []
    if search:
        where.append("(fc_code ILIKE %s OR city ILIKE %s OR state ILIKE %s OR fc_name ILIKE %s)")
        params.extend([f"%{search[:200]}%"] * 4)
    return Response(
        _paginated_select(
            table_sql="master.fc_master",
            columns=("fc_id", "fc_code", "fc_name", "city", "state", "region"),
            where=where,
            params=params,
            order_sql="ORDER BY fc_code ASC",
            page=page,
            page_size=page_size,
            offset=offset,
        )
    )


@api_view(["GET"])
@permission_classes([require("platform.po.view")])
def amazon_po_summary(request):
    _ensure_amazon_access(request.user)
    with connection.cursor() as cur:
        cur.execute(
            """
            SELECT
                COUNT(*) AS total_rows,
                COUNT(DISTINCT po_number) AS unique_pos,
                COUNT(DISTINCT fulfillment_center) AS unique_fcs,
                COUNT(*) FILTER (WHERE po_status IN ('MOV', 'PENDING')) AS mov_pending_count,
                COUNT(*) FILTER (WHERE
                    expiry_date IS NOT NULL
                    AND expiry_date >= CURRENT_DATE
                    AND expiry_date <= CURRENT_DATE + 7
                    AND po_status NOT IN ('COMPLETED', 'CANCELLED', 'EXPIRED')
                ) AS expiring_soon_count,
                COALESCE(SUM(total_requested_cost), 0) AS total_order_value,
                COALESCE(SUM(received_qty), 0) AS total_received_qty,
                COALESCE(SUM(requested_qty), 0) AS total_requested_qty,
                CASE WHEN SUM(requested_qty) > 0
                    THEN ROUND(100.0 * SUM(received_qty) / SUM(requested_qty), 1)
                    ELSE 0
                END AS fill_rate_pct
            FROM reporting."Amazon PO"
            """
        )
        row = cur.fetchone()
        cols = [desc[0] for desc in cur.description]
        summary = {col: _jsonable(val) for col, val in zip(cols, row)}

        cur.execute(
            """
            SELECT po_status, COUNT(*) AS count
            FROM reporting."Amazon PO"
            WHERE po_status IS NOT NULL
            GROUP BY po_status
            ORDER BY count DESC
            """
        )
        status_breakdown = [{"status": r[0], "count": int(r[1] or 0)} for r in cur.fetchall()]

        cur.execute(
            """
            SELECT category, COUNT(*) AS count,
                   COALESCE(SUM(total_requested_cost), 0) AS order_value,
                   CASE WHEN SUM(requested_qty) > 0
                       THEN ROUND(100.0 * SUM(received_qty) / SUM(requested_qty), 1)
                       ELSE 0
                   END AS fill_rate_pct
            FROM reporting."Amazon PO"
            WHERE category IS NOT NULL AND TRIM(category) != ''
            GROUP BY category
            ORDER BY count DESC
            LIMIT 8
            """
        )
        top_categories = _rows_to_dicts(cur)

        cur.execute(
            """
            SELECT fulfillment_center, COUNT(*) AS row_count,
                   CASE WHEN SUM(requested_qty) > 0
                       THEN ROUND(100.0 * SUM(received_qty) / SUM(requested_qty), 1)
                       ELSE 0
                   END AS fill_rate_pct,
                   COALESCE(SUM(total_requested_cost), 0) AS order_value
            FROM reporting."Amazon PO"
            WHERE fulfillment_center IS NOT NULL AND TRIM(fulfillment_center) != ''
            GROUP BY fulfillment_center
            ORDER BY row_count DESC
            LIMIT 10
            """
        )
        fc_breakdown = _rows_to_dicts(cur)

        # Item-head breakdown (all distinct item_head values)
        cur.execute(
            """
            SELECT COALESCE(NULLIF(TRIM(item_head), ''), 'Unknown') AS item_head,
                   COUNT(*) AS count,
                   COALESCE(SUM(total_requested_cost), 0) AS order_value,
                   CASE WHEN SUM(requested_qty) > 0
                       THEN ROUND(100.0 * SUM(received_qty) / SUM(requested_qty), 1)
                       ELSE 0
                   END AS fill_rate_pct
            FROM reporting."Amazon PO"
            GROUP BY COALESCE(NULLIF(TRIM(item_head), ''), 'Unknown')
            ORDER BY count DESC
            """
        )
        item_head_breakdown = _rows_to_dicts(cur)

        # State breakdown (top 12 states by row count)
        cur.execute(
            """
            SELECT COALESCE(NULLIF(TRIM(state), ''), 'Unknown') AS state,
                   COUNT(*) AS count,
                   COALESCE(SUM(total_requested_cost), 0) AS order_value,
                   CASE WHEN SUM(requested_qty) > 0
                       THEN ROUND(100.0 * SUM(received_qty) / SUM(requested_qty), 1)
                       ELSE 0
                   END AS fill_rate_pct
            FROM reporting."Amazon PO"
            WHERE state IS NOT NULL AND TRIM(state) != ''
            GROUP BY COALESCE(NULLIF(TRIM(state), ''), 'Unknown')
            ORDER BY count DESC
            LIMIT 12
            """
        )
        state_breakdown = _rows_to_dicts(cur)

        # Sub-category breakdown (top 10)
        cur.execute(
            """
            SELECT COALESCE(NULLIF(TRIM(sub_category), ''), 'Unknown') AS sub_category,
                   COUNT(*) AS count,
                   COALESCE(SUM(total_requested_cost), 0) AS order_value,
                   CASE WHEN SUM(requested_qty) > 0
                       THEN ROUND(100.0 * SUM(received_qty) / SUM(requested_qty), 1)
                       ELSE 0
                   END AS fill_rate_pct
            FROM reporting."Amazon PO"
            WHERE sub_category IS NOT NULL AND TRIM(sub_category) != ''
            GROUP BY COALESCE(NULLIF(TRIM(sub_category), ''), 'Unknown')
            ORDER BY count DESC
            LIMIT 10
            """
        )
        sub_category_breakdown = _rows_to_dicts(cur)

        # Classification KPIs: PREMIUM / COMMODITY / OTHERS by item_head
        cur.execute(
            """
            SELECT
                UPPER(TRIM(COALESCE(item_head, ''))) AS cls,
                COUNT(*) AS count,
                COALESCE(SUM(total_requested_cost), 0) AS order_value,
                CASE WHEN SUM(requested_qty) > 0
                    THEN ROUND(100.0 * SUM(received_qty) / SUM(requested_qty), 1)
                    ELSE 0
                END AS fill_rate_pct
            FROM reporting."Amazon PO"
            WHERE UPPER(TRIM(COALESCE(item_head, ''))) IN ('PREMIUM', 'COMMODITY', 'OTHERS')
            GROUP BY UPPER(TRIM(COALESCE(item_head, '')))
            """
        )
        classification_rows = {r[0]: {"count": int(r[1] or 0), "order_value": _jsonable(r[2]), "fill_rate_pct": _jsonable(r[3])} for r in cur.fetchall()}
        classification_kpis = {
            "premium":   classification_rows.get("PREMIUM",   {"count": 0, "order_value": 0, "fill_rate_pct": 0}),
            "commodity": classification_rows.get("COMMODITY", {"count": 0, "order_value": 0, "fill_rate_pct": 0}),
            "others":    classification_rows.get("OTHERS",    {"count": 0, "order_value": 0, "fill_rate_pct": 0}),
        }

        # Expiry urgency: active POs expiring within 7 days
        cur.execute(
            """
            SELECT po_number, fulfillment_center, state,
                   expiry_date, (expiry_date - CURRENT_DATE) AS days_left,
                   po_status,
                   CASE WHEN requested_qty > 0
                       THEN ROUND(100.0 * received_qty / requested_qty, 1)
                       ELSE 0
                   END AS fill_rate_pct,
                   total_requested_cost AS order_value
            FROM reporting."Amazon PO"
            WHERE expiry_date IS NOT NULL
              AND expiry_date >= CURRENT_DATE
              AND expiry_date <= CURRENT_DATE + 7
              AND po_status NOT IN ('COMPLETED', 'CANCELLED', 'EXPIRED')
            ORDER BY expiry_date ASC, po_number
            LIMIT 20
            """
        )
        expiry_urgent = _rows_to_dicts(cur)

    return Response(
        {
            "summary": summary,
            "status_breakdown": status_breakdown,
            "top_categories": top_categories,
            "fc_breakdown": fc_breakdown,
            "item_head_breakdown": item_head_breakdown,
            "state_breakdown": state_breakdown,
            "sub_category_breakdown": sub_category_breakdown,
            "classification_kpis": classification_kpis,
            "expiry_urgent": expiry_urgent,
        }
    )


@api_view(["GET"])
@permission_classes([require("platform.po.view")])
def appointment_summary(request):
    _ensure_amazon_access(request.user)
    with connection.cursor() as cur:
        cur.execute(
            """
            SELECT
                COUNT(*) AS total_rows,
                COUNT(DISTINCT appointment_id) AS unique_appointments,
                COUNT(DISTINCT destination_fc) AS unique_fcs,
                COUNT(*) FILTER (WHERE LOWER(status) = 'confirmed') AS confirmed_count,
                COUNT(*) FILTER (WHERE LOWER(status) = 'closed') AS closed_count,
                COUNT(*) FILTER (WHERE LOWER(status) = 'cancelled') AS cancelled_count,
                COUNT(*) FILTER (WHERE DATE(appointment_time) = CURRENT_DATE) AS today_count,
                COUNT(*) FILTER (WHERE
                    appointment_time >= date_trunc('week', CURRENT_DATE)
                    AND appointment_time < date_trunc('week', CURRENT_DATE) + interval '7 days'
                ) AS this_week_count
            FROM reporting."appointment"
            """
        )
        row = cur.fetchone()
        cols = [desc[0] for desc in cur.description]
        summary = {col: _jsonable(val) for col, val in zip(cols, row)}

        cur.execute(
            """
            SELECT status, COUNT(*) AS count
            FROM reporting."appointment"
            WHERE status IS NOT NULL AND TRIM(status) != ''
            GROUP BY status
            ORDER BY count DESC
            """
        )
        status_breakdown = [{"status": r[0], "count": int(r[1] or 0)} for r in cur.fetchall()]

        cur.execute(
            """
            SELECT destination_fc, COUNT(*) AS count,
                   COUNT(*) FILTER (WHERE LOWER(status) = 'confirmed') AS confirmed_count
            FROM reporting."appointment"
            WHERE destination_fc IS NOT NULL AND TRIM(destination_fc) != ''
            GROUP BY destination_fc
            ORDER BY count DESC
            LIMIT 10
            """
        )
        fc_breakdown = _rows_to_dicts(cur)

        cur.execute(
            """
            SELECT DATE(appointment_time) AS appt_date, COUNT(*) AS count
            FROM reporting."appointment"
            WHERE appointment_time >= CURRENT_DATE - interval '14 days'
              AND appointment_time < CURRENT_DATE + interval '7 days'
            GROUP BY DATE(appointment_time)
            ORDER BY appt_date ASC
            """
        )
        daily_counts = [
            {"date": str(r[0]), "count": int(r[1] or 0)} for r in cur.fetchall()
        ]

        # PRO breakdown (top 15 by count)
        cur.execute(
            """
            SELECT COALESCE(NULLIF(TRIM(pro), ''), 'Unknown') AS pro,
                   COUNT(*) AS count,
                   COUNT(*) FILTER (WHERE LOWER(status) = 'confirmed') AS confirmed_count,
                   COUNT(*) FILTER (WHERE LOWER(status) = 'closed') AS closed_count,
                   COUNT(*) FILTER (WHERE LOWER(status) = 'cancelled') AS cancelled_count
            FROM reporting."appointment"
            WHERE pro IS NOT NULL AND TRIM(pro) != ''
            GROUP BY COALESCE(NULLIF(TRIM(pro), ''), 'Unknown')
            ORDER BY count DESC
            LIMIT 15
            """
        )
        pro_breakdown = _rows_to_dicts(cur)

        # Month-over-month trend (current month vs previous month)
        cur.execute(
            """
            SELECT
                to_char(date_trunc('month', appointment_time), 'Mon YYYY') AS month_label,
                date_trunc('month', appointment_time) AS month_start,
                COUNT(*) AS total,
                COUNT(*) FILTER (WHERE LOWER(status) = 'confirmed') AS confirmed,
                COUNT(*) FILTER (WHERE LOWER(status) = 'closed') AS closed,
                COUNT(*) FILTER (WHERE LOWER(status) = 'cancelled') AS cancelled
            FROM reporting."appointment"
            WHERE appointment_time >= date_trunc('month', CURRENT_DATE) - interval '1 month'
              AND appointment_time < date_trunc('month', CURRENT_DATE) + interval '1 month'
            GROUP BY date_trunc('month', appointment_time), to_char(date_trunc('month', appointment_time), 'Mon YYYY')
            ORDER BY month_start
            """
        )
        mom_rows = cur.fetchall()
        mom_trend = []
        for r in mom_rows:
            mom_trend.append({
                "month_label": r[0],
                "total": int(r[2] or 0),
                "confirmed": int(r[3] or 0),
                "closed": int(r[4] or 0),
                "cancelled": int(r[5] or 0),
            })

    return Response(
        {
            "summary": summary,
            "status_breakdown": status_breakdown,
            "fc_breakdown": fc_breakdown,
            "daily_counts": daily_counts,
            "pro_breakdown": pro_breakdown,
            "mom_trend": mom_trend,
        }
    )
