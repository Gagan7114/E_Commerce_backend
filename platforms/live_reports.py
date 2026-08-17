"""
Live Reports (surfaced in the UI as "Pricing") — proxy + parse + cache for the
daily Excel workbooks published to the public GitHub repo
``daman8271/sending-excel``.

A bot commits ~12 dated ``.xlsx`` files there each day, named
``<Report>-YYYY-MM-DD.xlsx`` (competitor price watch + Jivo live reports across
Amazon / Fresh / Now, BigBasket, Blinkit, Flipkart, Flipkart Minutes, Price
Match). We surface them as a generic, multi-sheet grid viewer:

* ``GET /api/reports/live/reports`` — list reports (latest date per report prefix)
* ``GET /api/reports/live/data``    — one sheet as a raw grid (header + paginated,
  searchable body rows)

No DB tables are involved. The latest workbook is downloaded from GitHub, parsed
with openpyxl, and cached **content-addressed by the GitHub blob sha**, so a new
day's file (or a same-day re-commit that changes the sha) refreshes automatically
with no cron. Everything is fully generic: any ``*-YYYY-MM-DD.xlsx`` in the repo
shows up as a report on its own, no per-report code.

Cell values are rendered the way the workbook renders them: openpyxl is asked for
cell objects (not bare values) so each cell's ``number_format`` survives, and
``_cell`` applies it. Without this a ``0.0%`` cell holding ``-0.164`` reaches the
browser as ``-0.164`` instead of ``-16.4%``, and every ``"Rs"#,##0`` price loses
its currency mark — on a pricing screen that is a wrong number, not a cosmetic
one. Formats are read in READ-ONLY mode, which costs no extra memory (measured:
1.3 MB either way on the largest workbook; a non-read-only load costs 67x more).
It does cost ~31% more parse CPU (a day's set: 4.7s -> 6.1s), which the LRU below
absorbs — a workbook is parsed once and then answers every page and search from
memory.

Robustness notes:
* Downloads use stdlib ``urllib`` with a short timeout; every network / parse /
  decode failure is funnelled through ``_SourceError`` and returned as a 502 with
  a generic message (details are logged server-side, never echoed to the client).
* A size cap (before parsing) and a per-sheet row cap bound worker memory against
  an oversized or pathological upstream file.
* Parsed workbooks are held in a per-process LRU so repeated page/search requests
  for one report avoid both a re-download and a re-parse. The LRU is bounded by
  RETAINED CELL COUNT (not workbook count), because workbooks differ by two
  orders of magnitude: the nine reports published on 2026-08-17 total 634,972
  cells / 33.8 MB, but a single one of them ranges from 1,774 to 222,843 cells.
* The listing uses the Git Trees API (``/git/trees/{branch}?recursive=1``), which
  has no 1000-entry ceiling, and falls back to the Contents API if the tree call
  fails. NOTE: tree blobs carry no ``download_url``, so the raw URL is built here
  — it is used both to fetch the workbook and as the UI's "Download .xlsx" link.
  (The source repo is rolled daily rather than accumulating, so the old Contents
  API ceiling was never actually reached; the Trees API simply removes the cliff.)
"""
from __future__ import annotations

import datetime
import decimal
import http.client
import io
import json
import logging
import os
import re
import threading
import urllib.error
import urllib.request
from collections import OrderedDict
from typing import Any
from urllib.parse import quote

from django.conf import settings
from django.core.cache import cache
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from accounts.permissions import require

try:  # openpyxl is a hard dependency, but guard like the rest of the codebase does.
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

logger = logging.getLogger(__name__)


# --- Source config -----------------------------------------------------------
GITHUB_OWNER = "daman8271"
GITHUB_REPO = "sending-excel"
GITHUB_BRANCH = "main"
_CONTENTS_API = f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/"
_TREES_API = (
    f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}"
    f"/git/trees/{GITHUB_BRANCH}?recursive=1"
)
_RAW_BASE = f"https://raw.githubusercontent.com/{GITHUB_OWNER}/{GITHUB_REPO}/{GITHUB_BRANCH}/"

# Cache / limit knobs.
_LIST_TTL = 15 * 60           # GitHub directory listing — 15 min
_LIST_FAIL_TTL = 60           # back off this long after a failed listing (rate limits)
_FILE_TTL = 12 * 60 * 60      # downloaded workbook bytes — 12 h
_HTTP_TIMEOUT = 30            # seconds (per socket op) — short so a slow GitHub can't pin a worker
_MAX_BYTES = 30 * 1024 * 1024 # reject/parse-guard: workbooks larger than 30 MB
_MAX_ROWS = 50_000            # per-sheet row cap materialised into the grid
_DEFAULT_PAGE_SIZE = 100
_MAX_PAGE_SIZE = 500
_FULL_GRID_CAP = 2000        # rows returned to the Dashboard view (full=1), un-paginated

# Per-process LRU of parsed workbooks, keyed by blob sha (content-addressed).
# Avoids the pickle round-trip / re-parse a shared cache would incur on every page
# request. Bounded by retained CELLS rather than workbook count: measured at ~53
# bytes per cell, so 750k cells ~= 40 MB per worker, which holds an entire day's
# report set (634,972 cells on 2026-08-17) without evicting on every rail click.
# A workbook ceiling still applies as a belt-and-braces bound.
_PARSE_LRU_MAX = 12
_PARSE_LRU_MAX_CELLS = 750_000
_PARSE_LRU: "OrderedDict[str, dict]" = OrderedDict()
_PARSE_LRU_CELLS = 0
_LRU_LOCK = threading.Lock()

# ``<report prefix>-YYYY-MM-DD.xlsx``
_DATED_RE = re.compile(r"^(?P<prefix>.+)-(?P<date>\d{4}-\d{2}-\d{2})\.xlsx$", re.IGNORECASE)


class _SourceError(Exception):
    """Any failure reaching, decoding, or parsing the upstream reports source.
    Views translate this into a 502 with a generic client message."""


# --- HTTP --------------------------------------------------------------------
def _github_token() -> str:
    """Optional token for when the repo goes private. Empty ⇒ unauthenticated
    (fine while the repo is public). To enable, add ``GITHUB_REPORTS_TOKEN`` to
    the backend ``.env``; django-environ loads it into the process environment."""
    return getattr(settings, "GITHUB_REPORTS_TOKEN", "") or os.environ.get("GITHUB_REPORTS_TOKEN", "")


def _http_get(url: str) -> bytes:
    headers = {
        "User-Agent": "ecms-live-reports",
        "Accept": "application/vnd.github+json",
    }
    token = _github_token()
    if token:
        headers["Authorization"] = f"Bearer {token}"
    req = urllib.request.Request(url, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=_HTTP_TIMEOUT) as resp:
            return resp.read()
    except (OSError, http.client.HTTPException) as e:
        # OSError covers URLError/HTTPError (URLError subclasses OSError) and
        # read-phase socket.timeout/TimeoutError; HTTPException covers IncompleteRead.
        raise _SourceError("Could not reach the reports source.") from e


# --- Repo listing ------------------------------------------------------------
def _decode_json(raw: bytes):
    try:
        return json.loads(raw.decode("utf-8"))
    except (ValueError, UnicodeDecodeError) as e:
        raise _SourceError("The reports source returned an unreadable listing.") from e


def _list_via_trees() -> list[dict]:
    """Repo-root files via the Git Trees API — no 1000-entry ceiling.

    Tree blobs carry no ``download_url``, so we build the raw URL ourselves; it is
    used both to fetch the workbook server-side and as the UI's download link.
    Only root-level blobs are returned, matching what the Contents API listed.
    """
    data = _decode_json(_http_get(_TREES_API))
    if not isinstance(data, dict) or not isinstance(data.get("tree"), list):
        raise _SourceError("Unexpected response from the reports source.")
    if data.get("truncated"):
        # Genuinely enormous repo. Log it — this is the one case the Trees API
        # cannot serve in a single call, and it must not fail silently.
        logger.warning("live_reports: git tree response was truncated by GitHub")
    files = []
    for node in data["tree"]:
        if not isinstance(node, dict) or node.get("type") != "blob":
            continue
        path = node.get("path") or ""
        if not isinstance(path, str) or not path or "/" in path:  # root level only
            continue
        files.append(
            {
                "name": path,
                "type": "file",
                "sha": node.get("sha"),
                "size": node.get("size"),
                "download_url": _RAW_BASE + quote(path),
            }
        )
    return files


def _list_via_contents() -> list[dict]:
    """Fallback listing. Capped at 1000 entries by GitHub — fine as a backstop."""
    data = _decode_json(_http_get(_CONTENTS_API))
    if not isinstance(data, list):
        raise _SourceError("Unexpected response from the reports source.")
    return [f for f in data if isinstance(f, dict) and f.get("type") == "file"]


def _list_repo_files() -> list[dict]:
    """Repo-root file listing (cached ~15 min). Trees API first, Contents as fallback."""
    cached = cache.get("livereports.contents")
    if cached is not None:
        return cached
    # Unauthenticated GitHub allows 60 requests/hour per server IP. Once that is
    # spent, every page load would otherwise retry and dig the hole deeper, so a
    # failure is briefly negative-cached. Short enough to recover on its own.
    if cache.get("livereports.contents.failed"):
        raise _SourceError("The reports source is rate-limited or unreachable.")
    try:
        try:
            files = _list_via_trees()
        except _SourceError as e:
            logger.warning("live_reports: trees listing failed (%s); falling back to contents",
                           e.__cause__ or e)
            files = _list_via_contents()  # raises _SourceError if this fails too
    except _SourceError:
        cache.set("livereports.contents.failed", True, timeout=_LIST_FAIL_TTL)
        raise
    cache.set("livereports.contents", files, timeout=_LIST_TTL)
    return files


def _pretty_label(prefix: str) -> str:
    """``Jivo-AmazonFresh-Live-Report`` → ``Jivo AmazonFresh Live Report``."""
    return re.sub(r"[-_]+", " ", prefix).strip()


def _latest_reports() -> list[dict]:
    """Group dated ``.xlsx`` files by report prefix and keep the newest date each."""
    latest: dict[str, dict] = {}
    for f in _list_repo_files():
        name = f.get("name", "")
        m = _DATED_RE.match(name)
        if not m:
            continue
        prefix, date = m.group("prefix"), m.group("date")
        # The regex only proves the SHAPE is YYYY-MM-DD. Reject impossible dates:
        # "latest" is a string comparison, so `Report-2026-13-45.xlsx` would sort
        # above every real date and hijack the report.
        try:
            parsed_date = datetime.date.fromisoformat(date)
        except ValueError:
            logger.warning("live_reports: ignoring %r — %r is not a real date", name, date)
            continue
        # A far-future date is just as effective a hijack as an impossible one:
        # `Report-9999-12-31.xlsx` is a valid date that outsorts every real file.
        if parsed_date > datetime.date.today() + datetime.timedelta(days=2):
            logger.warning("live_reports: ignoring %r — %r is implausibly far ahead", name, date)
            continue
        cur = latest.get(prefix)
        if cur is None or date > cur["date"]:
            latest[prefix] = {
                "key": prefix,                      # stable key = filename prefix
                "label": _pretty_label(prefix),
                "date": date,
                "filename": name,
                "download_url": f.get("download_url"),
                "sha": f.get("sha"),                # content id — cache is keyed on this
                "size": f.get("size"),
            }
    return sorted(latest.values(), key=lambda r: r["label"].lower())


def _find_report(key: str) -> dict | None:
    for r in _latest_reports():
        if r["key"] == key:
            return r
    return None


def _last_commit_iso(report: dict) -> str | None:
    """ISO datetime (UTC) of the commit that last published this file — i.e. when
    the report was last updated on GitHub. Cached by sha. Best-effort: any failure
    returns None and the UI simply omits the time.

    NOTE: the source bot publishes all reports together in one daily commit, so
    this value is the same across every report/sheet on a given day; it changes
    day to day, not per sheet."""
    sha = report.get("sha") or report["filename"]
    ck = f"livereports.commit:{sha}"
    cached = cache.get(ck)
    if cached is not None:
        return cached or None
    url = (
        f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/commits"
        f"?path={quote(report['filename'])}&per_page=1"
    )
    iso = ""
    try:
        data = json.loads(_http_get(url).decode("utf-8"))
        # Guard the element type too: a well-formed-but-unexpected body like
        # `[null]` would otherwise raise AttributeError and surface as a 500
        # instead of the generic 502 this module promises.
        if isinstance(data, list) and data and isinstance(data[0], dict):
            iso = ((data[0].get("commit") or {}).get("committer") or {}).get("date") or ""
    except (_SourceError, ValueError, UnicodeDecodeError, KeyError, IndexError, TypeError,
            AttributeError):
        iso = ""
    cache.set(ck, iso, timeout=_FILE_TTL)
    return iso or None


# NOTE ON A TIMESTAMP THAT IS DELIBERATELY ABSENT FROM THE LIST ENDPOINT.
# It would be convenient to stamp every row of /reports with one "last updated"
# value, and the obvious cheap source is the branch head (one API call). That was
# tried and removed: the reports are NOT published in a single daily commit.
# Measured on 2026-08-17, 7 of 9 files differ from the branch head —
#   AllQcomm 05:30:02Z, Amazon-Now 05:00:02Z, AmazonFresh 04:30:02Z, ...
# so the branch head would have mislabelled most reports, and on a day when one
# scraper fails it would stamp a stale file with today's time — the worst case,
# because it hides exactly the staleness the field exists to reveal.
# Per-file accuracy costs one request per report, which is not affordable against
# an unauthenticated 60/hr quota. So /reports ships the file's `date` only, and
# the accurate per-file `updated_at` arrives with /data via `_last_commit_iso`.


# --- Number formats ----------------------------------------------------------
# Enough of Excel's format grammar to render what these workbooks actually use:
# "General", '"Rs"#,##0', '0.0%', '#,##0', '0.00', and the Indian lakh/crore form
# '[>=10000000]"₹"##\,##\,##\,##0;[>=100000]"₹"##\,##\,##0;"₹"#,##0'.
# Anything unrecognised falls through to the plain rendering — never raises.

_FMT_CONDITION_RE = re.compile(r"\[(>=|<=|<>|>|<|=)\s*(-?[\d.]+)\]")
_FMT_BRACKET_RE = re.compile(r"\[[^\]]*\]")
_FMT_LITERAL_RE = re.compile(r'"([^"]*)"')
_INDIAN_GROUPING_RE = re.compile(r"#,##,#|##\\,##")


def _split_format_sections(fmt: str) -> list[str]:
    """Split on ';' while ignoring separators inside quotes or [] blocks."""
    out, buf, in_quote, depth = [], [], False, 0
    for ch in fmt:
        if ch == '"':
            in_quote = not in_quote
        elif ch == "[" and not in_quote:
            depth += 1
        elif ch == "]" and not in_quote:
            depth = max(0, depth - 1)
        elif ch == ";" and not in_quote and depth == 0:
            out.append("".join(buf))
            buf = []
            continue
        buf.append(ch)
    out.append("".join(buf))
    return out


def _tokenize_format(section: str) -> list[tuple[str, bool]]:
    """Split a format section into (text, is_literal) tokens.

    Quoted runs and backslash escapes are literal text; everything else is
    pattern. Knowing which is which is what lets a literal be placed on the
    correct SIDE of the number — '"Rs"#,##0' prefixes, '#,##0" kg"' suffixes.
    """
    tokens: list[tuple[str, bool]] = []
    buf, in_quote, i = "", False, 0
    while i < len(section):
        ch = section[i]
        if ch == '"':
            if buf:
                tokens.append((buf, in_quote))
                buf = ""
            in_quote = not in_quote
        elif ch == "\\" and i + 1 < len(section):
            if buf:
                tokens.append((buf, in_quote))
                buf = ""
            tokens.append((section[i + 1], True))
            i += 2
            continue
        else:
            buf += ch
        i += 1
    if buf:
        tokens.append((buf, in_quote))
    return tokens


def _pick_format_section(fmt: str, value: float) -> str:
    """Choose the section of a multi-part format that applies to `value`.

    Conditional formats (``[>=100000]...;...``) win by their own test. Otherwise
    Excel's positional rule applies: positive; negative; zero.
    """
    sections = _split_format_sections(fmt)
    conditional = [s for s in sections if _FMT_CONDITION_RE.search(s)]
    if conditional:
        for sec in sections:
            m = _FMT_CONDITION_RE.search(sec)
            if m is None:
                return sec  # the unconditional "everything else" section
            op, raw = m.group(1), m.group(2)
            try:
                threshold = float(raw)
            except ValueError:
                continue
            if (
                (op == ">=" and value >= threshold)
                or (op == ">" and value > threshold)
                or (op == "<=" and value <= threshold)
                or (op == "<" and value < threshold)
                or (op == "=" and value == threshold)
                or (op == "<>" and value != threshold)
            ):
                return sec
        return sections[-1]
    if len(sections) >= 2 and value < 0:
        return sections[1]
    if len(sections) >= 3 and value == 0:
        return sections[2]
    return sections[0]


def _group_indian(digits: str) -> str:
    """1234567 -> 12,34,567 (last three, then pairs)."""
    if len(digits) <= 3:
        return digits
    head, tail = digits[:-3], digits[-3:]
    parts = []
    while len(head) > 2:
        parts.insert(0, head[-2:])
        head = head[:-2]
    if head:
        parts.insert(0, head)
    return ",".join(parts + [tail])


def _decimals_in(pattern: str) -> int:
    """Count the decimal placeholders after the last '.' in a numeric pattern."""
    if "." not in pattern:
        return 0
    tail = pattern.rsplit(".", 1)[1]
    return sum(1 for ch in tail if ch in "0#")


def _format_number(value: float, fmt: str) -> str | None:
    """Render `value` using the Excel number format `fmt`.

    Returns None when the format carries no numeric instruction we understand,
    so the caller can fall back to its plain rendering.
    """
    section = _pick_format_section(fmt, value)

    # A locale block ([$₹-en-IN]) carries a currency; its own leading '$' is
    # syntax, not a dollar sign, so it must be pulled out before any symbol scan.
    bracket_currency = ""
    for block in _FMT_BRACKET_RE.findall(section):
        if block.startswith("[$"):
            bracket_currency = block[2:-1].split("-", 1)[0]

    # Split the section at the numeric run so literals land on the side the
    # workbook put them: '"Rs"#,##0' prefixes, '#,##0" kg"' suffixes.
    body = _FMT_BRACKET_RE.sub("", section)
    is_percent = "%" in _FMT_LITERAL_RE.sub("", body)
    indian = bool(_INDIAN_GROUPING_RE.search(body))
    if indian:
        # In '##\,##\,##0' the escaped commas ARE the grouping separators, not
        # trailing literal text — unescape them so they stay part of the pattern.
        body = body.replace("\\,", ",")

    head, tail, skeleton_parts, seen_digit = [], [], [], False
    for token, is_literal in _tokenize_format(body):
        if is_literal:
            (tail if seen_digit else head).append(token)
            continue
        run = ""
        for ch in token:
            if ch in "0#?.,":
                seen_digit = True
                run += ch
            elif ch in "%_*":
                continue
            elif not seen_digit:
                head.append(ch)
            else:
                tail.append(ch)
        skeleton_parts.append(run)
    skeleton = "".join(skeleton_parts)

    if not any(ch in skeleton for ch in "0#?"):
        return None  # e.g. "General", or a pure text/date format

    head_str = "".join(head)
    suffix = "".join(tail)
    # The sign is re-applied below from the value itself, so a sign spelled in the
    # pattern must not also be pasted in literally.
    prefix = bracket_currency + head_str.replace("+", "").replace("-", "")
    # Accounting sections spell their own parentheses, and those DO flow through
    # as prefix/suffix — so the negative branch must not wrap them a second time.
    spells_parens = "(" in head_str and ")" in suffix
    # Excel spells an explicit positive sign in the pattern; dropping it erases
    # the above/below-reference signal on a delta column (and the UI colours on it).
    plus_sign = "+" if "+" in head_str and value > 0 else ""

    grouped = "," in skeleton or indian
    decimals = _decimals_in(skeleton)

    n = value * 100 if is_percent else value
    if n != n or n in (float("inf"), float("-inf")):
        return None

    negative = n < 0
    n = abs(n)
    # Excel rounds half AWAY FROM ZERO; Python's format() rounds half to even.
    # Without this, 380.5 under '"Rs"#,##0' renders Rs380 where the workbook
    # shows Rs381 — 313 real cells in today's set land on exactly that boundary.
    body = str(
        decimal.Decimal(repr(n)).quantize(
            decimal.Decimal(1).scaleb(-decimals), rounding=decimal.ROUND_HALF_UP
        )
    )
    int_part, _, frac_part = body.partition(".")
    if grouped:
        int_part = _group_indian(int_part) if indian else f"{int(int_part):,}"
    text = int_part + ("." + frac_part if frac_part else "")

    out = f"{prefix}{text}"
    if is_percent:
        out += "%"
    out += suffix
    if negative and not spells_parens:
        # Everything that does not spell parentheses takes a leading minus —
        # including a section written "-#,##0", which is spelling this same minus.
        out = "-" + out
    elif plus_sign:
        out = plus_sign + out
    return out


# --- Parsing -----------------------------------------------------------------
def _plain_number(v: float | int) -> str:
    if isinstance(v, int):
        return str(v)
    if v != v:  # NaN
        return ""
    if v.is_integer():
        return str(int(v))
    # Keep small magnitudes from collapsing to "0" WITHOUT falling back to
    # scientific notation, which has no place in a price grid.
    text = f"{v:.4f}".rstrip("0").rstrip(".")
    if text in ("", "0", "-0") and v != 0:
        text = f"{v:.12f}".rstrip("0").rstrip(".")
    if text in ("", "0", "-0") and v != 0:
        # Below 1e-12 even the widened decimal reads as zero. Showing a non-zero
        # value as "0" is worse than showing it in scientific notation.
        text = repr(v)
    return text


def _cell(v: Any, number_format: str | None = None) -> str:
    """Stringify a cell for display, honouring the workbook's number format.

    ``number_format`` is what makes a 0.0%-formatted -0.164 read as "-16.4%"
    instead of "-0.164". Formatting never raises: any unknown format falls back
    to the plain rendering.
    """
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)) and number_format and number_format != "General":
        try:
            formatted = _format_number(float(v), number_format)
        except Exception:  # a hostile format string must never break a report
            formatted = None
        if formatted is not None:
            return formatted
    if isinstance(v, (int, float)):
        return _plain_number(v)
    if isinstance(v, datetime.datetime):
        if (v.hour, v.minute, v.second) == (0, 0, 0):
            return v.strftime("%Y-%m-%d")
        return v.strftime("%Y-%m-%d %H:%M")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.time):
        return v.strftime("%H:%M")
    return str(v)


def _trim(grid: list[list[str]]) -> list[list[str]]:
    """Drop trailing all-empty rows/columns and pad every row to a uniform width
    (openpyxl's max_col/max_row over-report; read-only rows can be ragged)."""
    while grid and all(c == "" for c in grid[-1]):
        grid.pop()
    if not grid:
        return []
    width = 0
    for r in grid:
        last = 0
        for i, c in enumerate(r):
            if c != "":
                last = i + 1
        if last > width:
            width = last
    if width == 0:
        return []
    return [(r + [""] * width)[:width] for r in grid]


_SEARCH_STRIP_RE = re.compile(r"[,\s₹$€£]|rs\.?", re.IGNORECASE)


def _search_norm(text: str) -> str:
    """Strip grouping separators and currency marks for search comparison.

    "Rs1,499" and "1499" must both match a query of either form; without this,
    formatting the cells would have quietly broken numeric search.
    """
    return _SEARCH_STRIP_RE.sub("", (text or "").lower())


# A formatted cell that reads as a number: "489", "₹1,234.50", "Rs 489", "-16.4%",
# "12,34,567", "▲ 4.2%". Deliberately WIDER than the client's own isNumericish
# (it also accepts an "Rs" prefix, which these workbooks use heavily) — the client
# takes `numeric_cols` from the response rather than re-deriving it, so this is
# the single source of truth for alignment.
_NUMERICISH_RE = re.compile(
    r"^[+\-▲▼]?\s*(?:₹|Rs\.?|\$|€|£)?\s*-?\d[\d,]*(?:\.\d+)?\s*(?:%|L|ml|g|kg|x|X)?$",
    re.IGNORECASE,
)


def _numeric_columns(grid: list[list[str]], threshold: float = 0.7) -> list[bool]:
    """Decide, ONCE PER COLUMN over the whole sheet, whether it reads as numeric.

    The client used to infer this from whichever page was on screen, so a column
    could flip between left- and right-aligned as the user paged. Alignment is a
    property of the column, so it is settled here and sent with the response.
    """
    body = grid[1:] if len(grid) > 1 else []
    if not body:
        return []
    width = max((len(r) for r in grid), default=0)
    flags = []
    for c in range(width):
        filled = numeric = 0
        for row in body:
            cell = (row[c] if c < len(row) else "").strip()
            if not cell:
                continue
            filled += 1
            if _NUMERICISH_RE.match(cell):
                numeric += 1
        flags.append(filled > 0 and numeric / filled >= threshold)
    return flags


def _lru_get(sha: str) -> dict | None:
    with _LRU_LOCK:
        parsed = _PARSE_LRU.get(sha)
        if parsed is not None:
            _PARSE_LRU.move_to_end(sha)
        return parsed


def _lru_put(sha: str, parsed: dict) -> None:
    """Insert and evict oldest-first until both bounds hold (cells and count)."""
    global _PARSE_LRU_CELLS
    with _LRU_LOCK:
        prev = _PARSE_LRU.pop(sha, None)
        if prev is not None:
            _PARSE_LRU_CELLS -= prev.get("cells", 0)
        _PARSE_LRU[sha] = parsed
        _PARSE_LRU_CELLS += parsed.get("cells", 0)
        _PARSE_LRU.move_to_end(sha)
        while _PARSE_LRU and (
            len(_PARSE_LRU) > _PARSE_LRU_MAX or _PARSE_LRU_CELLS > _PARSE_LRU_MAX_CELLS
        ):
            if len(_PARSE_LRU) == 1:
                break  # never evict the entry just requested, however large
            _, dropped = _PARSE_LRU.popitem(last=False)
            _PARSE_LRU_CELLS -= dropped.get("cells", 0)


def _parse_bytes(content: bytes, report: dict) -> dict:
    if load_workbook is None:
        raise _SourceError("openpyxl is required to read report workbooks.")
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:  # BadZipFile / InvalidFileException / KeyError / …
        raise _SourceError("Could not parse the report workbook.") from e
    try:
        sheets = list(wb.sheetnames)
        grids: dict[str, list[list[str]]] = {}
        truncated_sheets: list[str] = []
        for name in sheets:
            ws = wb[name]
            grid: list[list[str]] = []
            hit_cap = False
            # Cell objects rather than values_only=True: this is what keeps each
            # cell's number_format, and in read-only mode it costs no extra memory.
            for i, row in enumerate(ws.iter_rows()):
                if i >= _MAX_ROWS:  # bound memory against a pathological sheet
                    hit_cap = True
                    break
                grid.append([_cell(c.value, getattr(c, "number_format", None)) for c in row])
            trimmed = _trim(grid)
            # Only a sheet that still HAS _MAX_ROWS of real content was truncated.
            # openpyxl over-reports max_row whenever a far-down cell carries only
            # styling, so testing the raw iteration count would raise a "large
            # sheet" banner over an 11-row sheet.
            if hit_cap and len(trimmed) >= _MAX_ROWS:
                truncated_sheets.append(name)
            grids[name] = trimmed
    finally:
        wb.close()
    cells = sum(len(r) for g in grids.values() for r in g)
    return {
        "sheets": sheets,
        "grids": grids,
        "date": report["date"],
        "filename": report["filename"],
        "truncated_sheets": truncated_sheets,
        "cells": cells,
        # Computed once per sheet at parse time and carried in the LRU. Doing it
        # per request cost 86 ms on the largest sheet, on every page click and
        # every keystroke of search.
        "numeric_cols": {n: _numeric_columns(g) for n, g in grids.items()},
    }


def _parse_workbook(report: dict) -> dict:
    """Return ``{sheets, grids, date, filename}`` for the report's latest workbook.

    Content-addressed by blob sha: a hit in the per-process LRU returns instantly;
    otherwise the raw bytes (cached ~12 h to avoid re-download) are parsed and the
    result is memoised in the LRU."""
    sha = report.get("sha") or report["filename"]

    hit = _lru_get(sha)
    if hit is not None:
        return hit

    size = report.get("size") or 0
    if size and size > _MAX_BYTES:
        raise _SourceError("The report file is too large to display.")

    bkey = f"livereports.bytes:{sha}"
    content = cache.get(bkey)
    if content is None:
        content = _http_get(report["download_url"])
        if len(content) > _MAX_BYTES:
            raise _SourceError("The report file is too large to display.")
        cache.set(bkey, content, timeout=_FILE_TTL)

    parsed = _parse_bytes(content, report)
    _lru_put(sha, parsed)
    return parsed


# --- Request helpers ---------------------------------------------------------
def _page_params(request) -> tuple[int, int]:
    def _int(name: str, default: int) -> int:
        try:
            return int(request.query_params.get(name, default))
        except (TypeError, ValueError):
            return default

    page = max(0, _int("page", 0))
    page_size = min(max(1, _int("page_size", _DEFAULT_PAGE_SIZE)), _MAX_PAGE_SIZE)
    return page, page_size


# --- Views -------------------------------------------------------------------
# Permission code that gates the whole Pricing section. Granted only to selected
# users (plus superusers) — see accounts/migrations/0017. A user who can't see
# the section can't reach its data either.
PRICING_PERMISSION = "pricing.view"
CanViewPricing = require(PRICING_PERMISSION)


@api_view(["GET"])
@permission_classes([IsAuthenticated, CanViewPricing])
def live_reports(request):
    """List the available reports (latest dated file per report)."""
    try:
        reports = _latest_reports()
    except _SourceError as e:
        logger.warning("live_reports source error: %s", e.__cause__ or e)
        return Response({"detail": "The reports source is currently unavailable."}, status=502)
    return Response(
        {
            "reports": [
                {
                    "key": r["key"],
                    "label": r["label"],
                    "date": r["date"],
                    "filename": r["filename"],
                    "download_url": r["download_url"],
                    "size": r["size"],
                }
                for r in reports
            ],
            "count": len(reports),
        }
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated, CanViewPricing])
def live_data(request):
    """Return one sheet of a report as a raw grid.

    Query params: ``report`` (key, required), ``sheet`` (defaults to first),
    ``q`` (row search), ``page`` (0-indexed), ``page_size``.

    Row 0 of the sheet is returned as ``header``; the remaining rows are the
    searchable, paginated ``rows`` body.
    """
    key = (request.query_params.get("report") or "").strip()
    if not key:
        return Response({"detail": "A 'report' query param is required."}, status=400)

    try:
        report = _find_report(key)
        if not report:
            return Response({"detail": "Unknown or unavailable report."}, status=404)
        parsed = _parse_workbook(report)
    except _SourceError as e:
        logger.warning("live_data source error (report=%s): %s", key, e.__cause__ or e)
        return Response({"detail": "The reports source is currently unavailable."}, status=502)

    sheets = parsed["sheets"]
    if not sheets:
        return Response(
            {
                "report": report["key"], "label": report["label"], "date": report["date"],
                "updated_at": _last_commit_iso(report),
                "download_url": report["download_url"], "sheets": [], "sheet": None,
                "header": [], "rows": [], "count": 0, "page": 0, "page_size": _DEFAULT_PAGE_SIZE,
                "rows_total": 0, "truncated": False,
                # `grid` must be present even here: the Sheet view keys its loading
                # state off Array.isArray(data.grid), so omitting it spins forever
                # on a workbook that legitimately has no sheets.
                "grid": [], "numeric_cols": [], "query": "",
                "full_grid_cap": _FULL_GRID_CAP, "row_cap_hit": False,
            }
        )

    requested = (request.query_params.get("sheet") or "").strip()
    sheet = requested if requested in parsed["grids"] else sheets[0]
    grid = parsed["grids"].get(sheet, [])

    # Dashboard view: return the whole sheet (un-paginated, capped) so the client
    # can interpret its structure into KPI cards / charts / tables. The Table view
    # keeps using the paginated body below.
    if str(request.query_params.get("full") or "").lower() in ("1", "true", "yes"):
        return Response(
            {
                "report": report["key"],
                "label": report["label"],
                "date": report["date"],
                "updated_at": _last_commit_iso(report),
                "download_url": report["download_url"],
                "sheets": sheets,
                "sheet": sheet,
                "grid": grid[:_FULL_GRID_CAP],
                "rows_total": len(grid),
                "truncated": len(grid) > _FULL_GRID_CAP,
                # Sent so the UI can explain the cap precisely and route oversized
                # sheets to the paginated Table view instead of implying it has
                # shown everything.
                "full_grid_cap": _FULL_GRID_CAP,
                "row_cap_hit": sheet in parsed.get("truncated_sheets", []),
            }
        )

    header = grid[0] if grid else []
    body = grid[1:] if len(grid) > 1 else []
    rows_total = len(body)  # before search, so the UI can say "12 of 4,555"

    q = (request.query_params.get("q") or "").strip().lower()
    if q:
        # Cells are now rendered the way the workbook renders them, so a price
        # cell reads "Rs1,499". Matching only that text would mean typing 1499
        # finds nothing — so match the literal text OR a separator/currency
        # stripped form of both sides, and searching either way works.
        qn = _search_norm(q)
        body = [
            r
            for r in body
            if any(q in (c or "").lower() or (qn and qn in _search_norm(c)) for c in r)
        ]

    total = len(body)
    page, page_size = _page_params(request)
    start = page * page_size
    page_rows = body[start : start + page_size]

    return Response(
        {
            "report": report["key"],
            "label": report["label"],
            "date": report["date"],
            "updated_at": _last_commit_iso(report),  # ISO UTC publish time (or null)
            "download_url": report["download_url"],
            "sheets": sheets,
            "sheet": sheet,
            "header": header,
            "rows": page_rows,
            "count": total,
            "page": page,
            "page_size": page_size,
            # Numeric alignment must be a property of the COLUMN, not of whichever
            # page happens to be on screen, so it is decided here over the whole
            # sheet and sent once.
            "numeric_cols": parsed.get("numeric_cols", {}).get(sheet, []),
            # `_MAX_ROWS` truncation was previously invisible on this path, so a
            # capped sheet under-reported `count` with no way for the UI to say so.
            "rows_total": rows_total,
            "truncated": sheet in parsed.get("truncated_sheets", []),
            "query": q,
        }
    )
