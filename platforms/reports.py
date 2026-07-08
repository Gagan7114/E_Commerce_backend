"""Raw-rows reports endpoint.

Exposes a whitelisted set of database views as JSON for the global Reports page
(Frontend/src/pages/Reports.jsx). No formulas, no transformations - just
SELECT <cols> FROM <view> WHERE <filters> LIMIT N.
"""

import io
import re
from datetime import datetime
from decimal import Decimal

from django.db import connection
from django.http import HttpResponse
from openpyxl import Workbook
from rest_framework.decorators import api_view, permission_classes
from rest_framework.exceptions import ValidationError
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from . import reports_sap


REPORT_VIEW_CATALOG = {
    "all_platform_inventory": {
        "date_column": "inventory_date",
        "format_column": "format",
        "max_rows": 50000,
    },
    "master_po": {
        "date_column": "po_date",
        "format_column": "format",
        "max_rows": 50000,
    },
    "prim_master_po": {
        "date_column": "po_date",
        "date_expr": "public._pm_parse_date(\"po_date\")",
        "format_column": "format",
        "max_rows": 50000,
    },
    "SecMaster": {
        "date_column": "date",
        "format_column": "format",
        "max_rows": 50000,
    },
    "amazon_sec_range_master_view": {
        "date_column": "to_date",
        "format_column": None,
        "max_rows": 50000,
    },
    "amazon_sec_daily_master_view": {
        "date_column": "to_date",
        "format_column": None,
        "max_rows": 50000,
    },
    # Amazon MP (Marketplace GST MTR B2B) raw table. Every column is stored
    # verbatim as TEXT (see platforms/migrations/0034_amazon_mp_table.py), so
    # there is no ::date-castable column to filter on — date_column=None means a
    # selected date range is ignored and the full row set exports. The table is
    # MP-only, so there is no format column to filter either.
    "amazon_mp": {
        "date_column": None,
        "format_column": None,
        "max_rows": 50000,
    },
    # Amazon MP enriched master VIEW — amazon_mp joined to master_sheet for
    # item_head / category / brand / delivered litres. What the Reports page now
    # reads for Amazon MP. Its date columns are mixed free-text formats, so date
    # filtering stays off (date_column=None); MP-only, so no format column.
    "amazon_mp_master_view": {
        "date_column": None,
        "format_column": None,
        "max_rows": 50000,
    },
}

# Some catalog "views" have a faster physical backing to actually query from.
# `SecMaster` is a LIVE view that recomputes a per-row landing-rate lookup, so a
# COUNT(*) is ~11s and a full export ~6 min (times out behind the web server).
# `secmaster_mv` is its materialized twin with identical columns — COUNT is
# ~0.1s and a full scan ~30s — so we run the report's queries against it while
# keeping the catalog/columns keyed on `SecMaster`.
_PHYSICAL_VIEW = {
    "SecMaster": "secmaster_mv",
}

_IDENT = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")
_DATE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _safe_view(name: str) -> str:
    name = (name or "").strip()
    if name not in REPORT_VIEW_CATALOG:
        raise ValidationError(f"Unknown report view: {name!r}")
    return name


def _safe_col(name: str) -> str:
    if not name or not _IDENT.match(name):
        raise ValidationError(f"Invalid column name: {name!r}")
    return name


def _normalised_format(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").lower())


def _parse_formats(raw: str) -> list[str]:
    """Split a comma-separated platform filter into distinct normalised formats.

    The UI may send one platform ("BIG BASKET") or several ("BLINKIT,ZEPTO")
    when the user multi-selects. Blank/duplicate entries are dropped; order is
    preserved so the generated IN-clause params are stable."""
    seen: list[str] = []
    for part in (raw or "").split(","):
        norm = _normalised_format(part)
        if norm and norm not in seen:
            seen.append(norm)
    return seen


def _format_filter(col: str, formats: list[str]) -> tuple[str, list]:
    """Build a normalised platform WHERE fragment + params for one or many formats.

    A single format keeps the original `= %s` shape; multiple formats become an
    `IN (%s, %s, ...)` over the same REGEXP_REPLACE-normalised column expression."""
    expr = f"REGEXP_REPLACE(LOWER(TRIM(\"{col}\"::text)), '[^a-z0-9]+', '', 'g')"
    if len(formats) == 1:
        return f"{expr} = %s", [formats[0]]
    placeholders = ", ".join(["%s"] * len(formats))
    return f"{expr} IN ({placeholders})", list(formats)


# xlsx hard cap is 1,048,576 rows (incl. the header) — stay safely under it.
EXPORT_MAX_ROWS = 1_000_000


def _coerce(v):
    """Make a DB value safe for an openpyxl write-only cell."""
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, datetime) and v.tzinfo is not None:
        return v.replace(tzinfo=None)  # openpyxl can't store tz-aware datetimes
    if isinstance(v, (bytes, bytearray, memoryview)):
        return bytes(v).decode("utf-8", "replace")
    return v


def _build_filters(catalog, fmt, date_from, date_to):
    """WHERE clause + params from the platform/date filters (shared by raw+export).

    `fmt` may be a single platform or a comma-separated list (multi-select)."""
    where_parts: list[str] = []
    params: list = []
    formats = _parse_formats(fmt)
    if formats and catalog["format_column"]:
        frag, frag_params = _format_filter(catalog["format_column"], formats)
        where_parts.append(frag)
        params.extend(frag_params)
    if catalog["date_column"]:
        date_expr = catalog.get("date_expr") or f'("{catalog["date_column"]}")::date'
        if date_from:
            if not _DATE.match(date_from):
                raise ValidationError("`date_from` must be YYYY-MM-DD.")
            where_parts.append(f"{date_expr} >= %s")
            params.append(date_from)
        if date_to:
            if not _DATE.match(date_to):
                raise ValidationError("`date_to` must be YYYY-MM-DD.")
            where_parts.append(f"{date_expr} <= %s")
            params.append(date_to)
    where_clause = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""
    return where_clause, params


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def report_columns(request):
    view_raw = request.query_params.get("view", "")
    if reports_sap.is_sap_view(view_raw):
        try:
            sap_cols = reports_sap.columns_for(view_raw)
        except ValueError as exc:
            raise ValidationError(str(exc))
        return Response({
            "view": view_raw,
            "columns": [{"key": name, "type": dtype} for name, dtype in sap_cols],
        })
    view = _safe_view(view_raw)
    with connection.cursor() as cur:
        cur.execute(
            """
            SELECT column_name, data_type
            FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = %s
            ORDER BY ordinal_position
            """,
            [view],
        )
        cols = cur.fetchall()
    return Response({
        "view": view,
        "columns": [{"key": name, "type": dtype} for name, dtype in cols],
    })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def report_raw(request):
    view_raw = request.query_params.get("view", "")
    if reports_sap.is_sap_view(view_raw):
        try:
            page = max(0, int(request.query_params.get("page") or 0))
        except ValueError:
            page = 0
        try:
            page_size = int(request.query_params.get("page_size") or 200)
        except ValueError:
            page_size = 200
        page_size = max(1, min(50000, page_size))
        date_from = (request.query_params.get("date_from") or "").strip()
        date_to = (request.query_params.get("date_to") or "").strip()
        if date_from and not _DATE.match(date_from):
            raise ValidationError("`date_from` must be YYYY-MM-DD.")
        if date_to and not _DATE.match(date_to):
            raise ValidationError("`date_to` must be YYYY-MM-DD.")
        try:
            rows, count = reports_sap.fetch_for(
                view_raw,
                from_date=date_from,
                to_date=date_to,
                page=page,
                page_size=page_size,
            )
        except ValueError as exc:
            raise ValidationError(str(exc))
        except Exception as exc:  # noqa: BLE001 — surface SAP errors to the UI
            return Response({
                "view": view_raw,
                "rows": [],
                "columns": [],
                "count": 0,
                "page": page,
                "page_size": page_size,
                "error": str(exc),
            })
        response_columns = list(rows[0].keys()) if rows else []
        return Response({
            "view": view_raw,
            "rows": rows,
            "columns": response_columns,
            "count": count,
            "page": page,
            "page_size": page_size,
        })
    view = _safe_view(view_raw)
    catalog = REPORT_VIEW_CATALOG[view]
    physical = _PHYSICAL_VIEW.get(view, view)

    requested_columns = (request.query_params.get("columns") or "").strip()
    columns: list[str] = []
    if requested_columns:
        for c in requested_columns.split(","):
            c = c.strip()
            if not c:
                continue
            _safe_col(c)
            columns.append(c)
    select_clause = ", ".join(f'"{c}"' for c in columns) if columns else "*"

    where_parts: list[str] = []
    params: list = []

    fmt = (request.query_params.get("platform") or request.query_params.get("fmt") or "").strip()
    formats = _parse_formats(fmt)
    if formats and catalog["format_column"]:
        frag, frag_params = _format_filter(catalog["format_column"], formats)
        where_parts.append(frag)
        params.extend(frag_params)

    date_from = (request.query_params.get("date_from") or "").strip()
    date_to = (request.query_params.get("date_to") or "").strip()
    if catalog["date_column"]:
        date_expr = catalog.get("date_expr") or f'("{catalog["date_column"]}")::date'
        if date_from:
            if not _DATE.match(date_from):
                raise ValidationError("`date_from` must be YYYY-MM-DD.")
            where_parts.append(f"{date_expr} >= %s")
            params.append(date_from)
        if date_to:
            if not _DATE.match(date_to):
                raise ValidationError("`date_to` must be YYYY-MM-DD.")
            where_parts.append(f"{date_expr} <= %s")
            params.append(date_to)

    where_clause = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""

    try:
        page = max(0, int(request.query_params.get("page") or 0))
    except ValueError:
        page = 0
    try:
        page_size = int(request.query_params.get("page_size") or 200)
    except ValueError:
        page_size = 200
    page_size = max(1, min(catalog["max_rows"], page_size))
    offset = page * page_size

    # A deterministic ORDER BY is REQUIRED for correct LIMIT/OFFSET pagination.
    # Without it every page is a fresh, unordered scan of the (materialized) view,
    # so consecutive OFFSET windows can overlap — and if secmaster_mv is REFRESHed
    # between page requests the overlap is severe. A paging client (e.g. the
    # SecMaster Apps Script) then appends those overlapping windows and produces
    # exact-duplicate rows. Ordering by every selected column is stable across
    # refreshes (same source data -> same multiset -> same sequence) and works for
    # any view without per-view configuration.
    order_clause = ""
    if columns:
        order_clause = "ORDER BY " + ", ".join(str(i + 1) for i in range(len(columns)))

    try:
        with connection.cursor() as cur:
            cur.execute(f'SELECT COUNT(*) FROM "{physical}" {where_clause}', params)
            count_row = cur.fetchone()
            count = int(count_row[0]) if count_row else 0
            cur.execute(
                f'SELECT {select_clause} FROM "{physical}" {where_clause} {order_clause} LIMIT %s OFFSET %s',
                params + [page_size, offset],
            )
            description = cur.description or []
            response_columns = [c[0] for c in description]
            rows = [dict(zip(response_columns, row)) for row in cur.fetchall()]
    except Exception as exc:
        return Response({
            "view": view,
            "rows": [],
            "columns": [],
            "count": 0,
            "page": page,
            "page_size": page_size,
            "error": str(exc),
        })

    return Response({
        "view": view,
        "rows": rows,
        "columns": response_columns,
        "count": count,
        "page": page,
        "page_size": page_size,
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def report_export(request):
    """Stream ALL matching rows for the current report as one .xlsx file.

    The /raw endpoint caps at 50k for the on-screen table; this runs a single
    query and writes every row with openpyxl in write-only mode (bounded memory,
    chunked fetch), so big exports — hundreds of thousands of rows — come down as
    one Excel file. One query = one consistent snapshot, so there are no
    pagination gaps/duplicates. Body: {view, columns[], labels[], platform,
    date_from, date_to, filters[[k,v]...], filename}."""
    data = request.data or {}
    view_raw = (data.get("view") or "").strip()
    columns = [str(c).strip() for c in (data.get("columns") or []) if str(c).strip()]
    labels = [str(x) for x in (data.get("labels") or [])]
    date_from = (data.get("date_from") or "").strip()
    date_to = (data.get("date_to") or "").strip()
    if date_from and not _DATE.match(date_from):
        raise ValidationError("`date_from` must be YYYY-MM-DD.")
    if date_to and not _DATE.match(date_to):
        raise ValidationError("`date_to` must be YYYY-MM-DD.")

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Report")
    total = 0

    if reports_sap.is_sap_view(view_raw):
        try:
            sap_rows, _ = reports_sap.fetch_for(
                view_raw,
                from_date=date_from,
                to_date=date_to,
                page=0,
                page_size=EXPORT_MAX_ROWS,
            )
        except ValueError as exc:
            raise ValidationError(str(exc))
        keys = columns or (list(sap_rows[0].keys()) if sap_rows else [])
        ws.append(labels if labels and len(labels) == len(keys) else keys)
        for r in sap_rows[:EXPORT_MAX_ROWS]:
            ws.append([_coerce(r.get(k)) for k in keys])
            total += 1
    else:
        view = _safe_view(view_raw)
        catalog = REPORT_VIEW_CATALOG[view]
        physical = _PHYSICAL_VIEW.get(view, view)
        for c in columns:
            _safe_col(c)
        select_clause = ", ".join(f'"{c}"' for c in columns) if columns else "*"
        fmt = (data.get("platform") or data.get("fmt") or "").strip()
        where_clause, params = _build_filters(catalog, fmt, date_from, date_to)
        sql = (
            f'SELECT {select_clause} FROM "{physical}" {where_clause} '
            f"LIMIT {EXPORT_MAX_ROWS}"
        )
        with connection.cursor() as cur:
            cur.execute(sql, params)
            keys = columns or [c[0] for c in (cur.description or [])]
            ws.append(labels if labels and len(labels) == len(keys) else keys)
            while True:
                chunk = cur.fetchmany(2000)
                if not chunk:
                    break
                for row in chunk:
                    ws.append([_coerce(v) for v in row])
                    total += 1

    # Filters sheet — mirror the on-screen filters, then the real exported count.
    ws2 = wb.create_sheet("Filters")
    for pair in data.get("filters") or []:
        if isinstance(pair, (list, tuple)) and len(pair) >= 2:
            ws2.append([str(pair[0]), "" if pair[1] is None else str(pair[1])])
    ws2.append(["Total Rows", total])

    buf = io.BytesIO()
    wb.save(buf)
    raw_name = (data.get("filename") or f"report_{view_raw}").strip() or "report"
    filename = re.sub(r'[\\/:*?"<>|]+', "_", raw_name)
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"
    resp = HttpResponse(
        buf.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
